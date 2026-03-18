#!/usr/bin/env python3
"""Build an Excel MVP dashboard for ETF vs GDP growth comparison."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass

import pandas as pd
import yfinance as yf
from openpyxl.utils import get_column_letter

from build_combined_etf_weo import (
    COUNTRY_TO_ISO3,
    COUNTRY_TO_LCU,
    build_ticker_country_map,
    compute_annual_fx_quote_to_usd_returns,
    compute_annual_jan1_fx_returns,
    normalize_currency_code,
)


TIMEFRAME_ORDER = [
    "YTD",
    "1 Month",
    "3 Months",
    "6 Months",
    "1 Year",
    "3 Years",
    "5 Years",
    "10 Years",
    "MAX",
]

TIMEFRAME_SPECS: list[tuple[str, object]] = [
    ("YTD", "ytd"),
    ("1 Month", pd.DateOffset(months=1)),
    ("3 Months", pd.DateOffset(months=3)),
    ("6 Months", pd.DateOffset(months=6)),
    ("1 Year", pd.DateOffset(years=1)),
    ("3 Years", pd.DateOffset(years=3)),
    ("5 Years", pd.DateOffset(years=5)),
    ("10 Years", pd.DateOffset(years=10)),
    ("MAX", "max"),
]

ANNUAL_WINDOW_YEARS = 10
CAGR_HORIZONS = [1, 3, 5, 10]
COUNTRY_TO_REGION = {
    "Australia": "Oceania",
    "Austria": "Europe",
    "Belgium": "Europe",
    "Brazil": "Latin America",
    "Bulgaria": "Europe",
    "Canada": "North America",
    "China": "Asia",
    "France": "Europe",
    "Germany": "Europe",
    "Greece": "Europe",
    "Hong Kong": "Asia",
    "India": "Asia",
    "Indonesia": "Asia",
    "Italy": "Europe",
    "Japan": "Asia",
    "Kuwait": "Middle East",
    "Malaysia": "Asia",
    "Mexico": "Latin America",
    "Netherlands": "Europe",
    "Pakistan": "Asia",
    "Philippines": "Asia",
    "Poland": "Europe",
    "Saudi Arabia": "Middle East",
    "Singapore": "Asia",
    "South Africa": "Africa",
    "South Korea": "Asia",
    "Spain": "Europe",
    "Sweden": "Europe",
    "Switzerland": "Europe",
    "Taiwan": "Asia",
    "Thailand": "Asia",
    "Turkey": "Europe",
    "United Kingdom": "Europe",
    "United States": "North America",
    "Vietnam": "Asia",
}


@dataclass
class PricePoint:
    date: pd.Timestamp
    value: float


def last_on_or_before(series: pd.Series, target: pd.Timestamp) -> PricePoint | None:
    s = series.dropna()
    s = s[s.index <= target]
    if s.empty:
        return None
    idx = s.index[-1]
    return PricePoint(date=idx, value=float(s.loc[idx]))


def first_in_year(series: pd.Series, year: int) -> PricePoint | None:
    s = series.dropna()
    s = s[s.index.year == year]
    if s.empty:
        return None
    idx = s.index[0]
    return PricePoint(date=idx, value=float(s.iloc[0]))


def last_in_year(series: pd.Series, year: int) -> PricePoint | None:
    s = series.dropna()
    s = s[s.index.year == year]
    if s.empty:
        return None
    idx = s.index[-1]
    return PricePoint(date=idx, value=float(s.iloc[-1]))


def pct_return(start_value: float, end_value: float) -> float:
    return ((end_value / start_value) - 1.0) * 100.0


def cagr_from_total_return(total_return_pct: float, years: int) -> float:
    return (((1.0 + (total_return_pct / 100.0)) ** (1.0 / years)) - 1.0) * 100.0


def max_start_point(series: pd.Series) -> PricePoint | None:
    s = series.dropna()
    if s.empty:
        return None

    # Guard against bad launch prints / scale breaks near series start.
    ratio = (s / s.shift(1)).dropna()
    suspicious = ratio[(ratio < 0.2) | (ratio > 5.0)]
    if not suspicious.empty:
        window_end = s.index.min() + pd.DateOffset(days=120)
        early_breaks = suspicious[suspicious.index <= window_end]
        if not early_breaks.empty:
            s = s[s.index >= early_breaks.index.min()]
            if s.empty:
                return None

    idx = s.index[1] if len(s) > 1 else s.index[0]
    return PricePoint(date=idx, value=float(s.loc[idx]))


def choose_price_columns(columns: list[str]) -> dict[str, str]:
    """
    Return ticker -> price column.
    Prefers Adj Close if available, otherwise Close.
    """
    pattern = re.compile(r"^(.+?) - ([^- ]+) - (Adj Close|Close)$")
    ticker_fields: dict[str, dict[str, str]] = {}
    for col in columns:
        match = pattern.match(col)
        if not match:
            continue
        _, ticker, field = match.groups()
        ticker_fields.setdefault(ticker, {})[field] = col

    selected: dict[str, str] = {}
    for ticker, fields in ticker_fields.items():
        if "Adj Close" in fields:
            selected[ticker] = fields["Adj Close"]
        elif "Close" in fields:
            selected[ticker] = fields["Close"]
    return selected


def load_gdp_growth_maps(
    weo_csv: str,
) -> tuple[
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
    dict[tuple[str, int], float],
]:
    weo = pd.read_csv(weo_csv)
    weo["year"] = pd.to_numeric(weo["year"], errors="coerce").astype("Int64")
    weo["value"] = pd.to_numeric(weo["value"], errors="coerce")
    weo = weo.dropna(subset=["country_code", "year"])
    weo["year"] = weo["year"].astype(int)
    
    real_weo = weo[(weo["indicator"] == "NGDP_RPCH") & (weo["value"].notna())].copy()
    nominal_usd_weo = weo[(weo["indicator"] == "NGDPD_PCH") & (weo["value"].notna())].copy()
    nominal_lcu_weo = weo[(weo["indicator"] == "NGDP_PCH") & (weo["value"].notna())].copy()
    current_usd_weo = weo[(weo["indicator"] == "NGDPD") & (weo["value"].notna())].copy()
    current_lcu_weo = weo[(weo["indicator"] == "NGDP") & (weo["value"].notna())].copy()
    inflation_cpi_weo = weo[(weo["indicator"] == "PCPIPCH") & (weo["value"].notna())].copy()

    real_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in real_weo.itertuples(index=False)
    }
    nominal_usd_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in nominal_usd_weo.itertuples(index=False)
    }
    nominal_lcu_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in nominal_lcu_weo.itertuples(index=False)
    }
    current_usd_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in current_usd_weo.itertuples(index=False)
    }
    current_lcu_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in current_lcu_weo.itertuples(index=False)
    }
    inflation_cpi_map = {
        (str(r.country_code), int(r.year)): float(r.value)
        for r in inflation_cpi_weo.itertuples(index=False)
    }

    # Backward-compatible fallbacks if growth rows are missing.
    if not nominal_usd_map:
        gdp_levels = current_usd_weo.sort_values(["country_code", "year"])
        gdp_levels["nominal_growth"] = (
            gdp_levels.groupby("country_code")["value"].pct_change() * 100.0
        )
        nominal_usd_map = {
            (str(r.country_code), int(r.year)): float(r.nominal_growth)
            for r in gdp_levels.itertuples(index=False)
            if pd.notna(r.nominal_growth)
        }

    if not nominal_lcu_map:
        gdp_levels_lcu = current_lcu_weo.sort_values(["country_code", "year"])
        gdp_levels_lcu["nominal_growth"] = (
            gdp_levels_lcu.groupby("country_code")["value"].pct_change() * 100.0
        )
        nominal_lcu_map = {
            (str(r.country_code), int(r.year)): float(r.nominal_growth)
            for r in gdp_levels_lcu.itertuples(index=False)
            if pd.notna(r.nominal_growth)
        }

    country_fx_map: dict[tuple[str, int], float] = {}
    if not current_lcu_weo.empty and not current_usd_weo.empty:
        ngdp = current_lcu_weo[["country_code", "year", "value"]].rename(columns={"value": "ngdp_lcu"})
        ngdpd = current_usd_weo[["country_code", "year", "value"]].rename(columns={"value": "ngdp_usd"})
        fx_levels = ngdp.merge(ngdpd, on=["country_code", "year"], how="inner")
        if not fx_levels.empty:
            fx_levels["lcu_per_usd"] = fx_levels["ngdp_lcu"] / fx_levels["ngdp_usd"]
            fx_levels = fx_levels.sort_values(["country_code", "year"]).copy()
            fx_levels["country_lcu_vs_usd_weo_pct"] = (
                fx_levels.groupby("country_code")["lcu_per_usd"].shift(1) 
                / fx_levels["lcu_per_usd"] - 1.0
            ) * 100.0
            country_fx_map = {
                (str(r.country_code), int(r.year)): float(r.country_lcu_vs_usd_weo_pct)
                for r in fx_levels.itertuples(index=False)
                if pd.notna(r.country_lcu_vs_usd_weo_pct)
            }

    return real_map, nominal_lcu_map, nominal_usd_map, country_fx_map, current_usd_map, current_lcu_map, inflation_cpi_map


def gdp_cagr(
    gdp_growth_map: dict[tuple[str, int], float],
    country_code: str,
    end_year: int,
    years: int,
) -> float | None:
    start_year = end_year - years + 1
    finish_year = end_year
    values: list[float] = []
    for y in range(start_year, finish_year + 1):
        v = gdp_growth_map.get((country_code, y))
        if v is None:
            return None
        values.append(v)
    compounded = 1.0
    for v in values:
        compounded *= 1.0 + (v / 100.0)
    return ((compounded ** (1.0 / years)) - 1.0) * 100.0


def calculate_inflation_metrics(
    inflation_map: dict[tuple[str, int], float],
    country_code: str,
    end_year: int,
    years: int,
) -> tuple[float | None, float | None, float | None]:
    """Calculate local CAGR, USA CAGR, and the simple difference."""
    start_year = end_year - years + 1
    
    def get_cagr(cc):
        compounded = 1.0
        for y in range(start_year, end_year + 1):
            v = inflation_map.get((cc, y))
            if v is None:
                return None
            compounded *= (1.0 + v / 100.0)
        return ((compounded ** (1.0 / years)) - 1.0) * 100.0

    cagr_local = get_cagr(country_code)
    cagr_usa = get_cagr("USA")
    
    diff = (cagr_local - cagr_usa) if (cagr_local is not None and cagr_usa is not None) else None
    return cagr_local, cagr_usa, diff


def fx_level_cagr(
    lcu_map: dict[tuple[str, int], float],
    usd_map: dict[tuple[str, int], float],
    country_code: str,
    end_year: int,
    years: int,
) -> float | None:
    """Calculate CAGR of FX rate (USD per LCU) using levels."""
    start_year_for_base = end_year - years # Price at end of start_year_for_base
    
    lcu_start = lcu_map.get((country_code, start_year_for_base))
    usd_start = usd_map.get((country_code, start_year_for_base))
    lcu_end = lcu_map.get((country_code, end_year))
    usd_end = usd_map.get((country_code, end_year))
    
    if None in (lcu_start, usd_start, lcu_end, usd_end) or lcu_start == 0 or usd_end == 0:
        return None
    
    xr_start = lcu_start / usd_start # LCU per USD
    xr_end = lcu_end / usd_end
    
    # CAGR of USD per LCU = ( (1/XR_end) / (1/XR_start) )^(1/years) - 1
    # which is (XR_start / XR_end)^(1/years) - 1
    return ((xr_start / xr_end) ** (1.0 / years) - 1.0) * 100.0


def load_history_wide(etf_csv: str) -> pd.DataFrame:
    raw = pd.read_csv(etf_csv)
    raw["Date"] = pd.to_datetime(raw["Date"], errors="coerce")
    raw = raw.dropna(subset=["Date"]).sort_values("Date")
    cols = ["Date"] + [c for c in raw.columns if c.endswith(" - Close")]
    return raw[cols].copy()


def load_ticker_currency_map(metadata_csv: str | None) -> dict[str, str]:
    if not metadata_csv:
        return {}
    try:
        meta = pd.read_csv(metadata_csv)
    except FileNotFoundError:
        return {}
    required = {"ticker", "currency"}
    if not required.issubset(set(meta.columns)):
        return {}
    meta = meta.copy()
    meta["ticker"] = meta["ticker"].astype(str)
    meta["currency"] = meta["currency"].fillna("").astype(str)
    return (
        meta[["ticker", "currency"]]
        .drop_duplicates(subset=["ticker"], keep="first")
        .set_index("ticker")["currency"]
        .to_dict()
    )


def load_ticker_accumulating_map(metadata_csv: str | None) -> dict[str, str]:
    """Load ticker -> is_accumulating mapping from metadata."""
    if not metadata_csv:
        return {}
    try:
        meta = pd.read_csv(metadata_csv)
    except FileNotFoundError:
        return {}
    required = {"ticker", "is_accumulating"}
    if not required.issubset(set(meta.columns)):
        return {}
    meta = meta.copy()
    meta["ticker"] = meta["ticker"].astype(str)
    meta["is_accumulating"] = meta["is_accumulating"].fillna("").astype(str)
    return (
        meta[["ticker", "is_accumulating"]]
        .drop_duplicates(subset=["ticker"], keep="first")
        .set_index("ticker")["is_accumulating"]
        .to_dict()
    )


def load_ticker_exchange_map(metadata_csv: str | None) -> dict[str, str]:
    if not metadata_csv:
        return {}
    try:
        meta = pd.read_csv(metadata_csv)
    except FileNotFoundError:
        return {}
    required = {"ticker", "exchange"}
    if not required.issubset(set(meta.columns)):
        return {}
    exchange_labels = {
        "LSE": "London Stock Exchange",
        "PCX": "NYSE Arca",
        "NYQ": "NYSE",
        "ASE": "NYSE American",
        "NMS": "Nasdaq Global Select",
        "NGM": "Nasdaq Global Market",
        "NCM": "Nasdaq Capital Market",
        "NAS": "Nasdaq",
        "BTS": "Cboe BZX",
    }

    meta = meta.copy()
    meta["ticker"] = meta["ticker"].astype(str)
    meta["exchange"] = meta["exchange"].fillna("").astype(str)
    meta["exchange"] = meta["exchange"].map(lambda x: exchange_labels.get(x, x))
    return (
        meta[["ticker", "exchange"]]
        .drop_duplicates(subset=["ticker"], keep="first")
        .set_index("ticker")["exchange"]
        .to_dict()
    )


def load_ticker_fund_size_map(metadata_csv: str | None) -> dict[str, float]:
    if not metadata_csv:
        return {}
    try:
        meta = pd.read_csv(metadata_csv)
    except FileNotFoundError:
        return {}
    required = {"ticker", "fund_size"}
    if not required.issubset(set(meta.columns)):
        return {}
    meta = meta.copy()
    meta["ticker"] = meta["ticker"].astype(str)
    meta["fund_size"] = pd.to_numeric(meta["fund_size"], errors="coerce")
    out = (
        meta[["ticker", "fund_size"]]
        .drop_duplicates(subset=["ticker"], keep="first")
        .dropna(subset=["fund_size"])
        .set_index("ticker")["fund_size"]
        .to_dict()
    )
    return {k: float(v) for k, v in out.items()}


def fit_columns_to_header(ws, header_row: int) -> None:
    """Auto-size column widths based ONLY on the string length of the header row."""
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        width = len(str(value)) if value is not None else 0
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def fetch_daily_fx_series(currencies: set[str]) -> dict[str, pd.Series]:
    """Fetch daily FX history (CCY -> USD) for all relevant currencies."""
    out: dict[str, pd.Series] = {}
    for currency in sorted(currencies):
        ccy = normalize_currency_code(currency)
        if not ccy or ccy == "USD":
            continue

        series: pd.Series | None = None
        for pair, invert in ((f"{ccy}USD=X", False), (f"USD{ccy}=X", True)):
            try:
                tk = yf.Ticker(pair)
                hist = tk.history(period="max", interval="1d", auto_adjust=True, actions=False)
                if not hist.empty and "Close" in hist.columns:
                    s = hist["Close"].dropna()
                    if not s.empty:
                        if getattr(s.index, "tz", None) is not None:
                            s.index = s.index.tz_localize(None)
                        if invert:
                            s = 1.0 / s
                        series = s
                        break
            except Exception:
                continue
        if series is not None:
            out[ccy] = series
    return out


def get_spot_fx_rates(currencies: set[str]) -> dict[str, float]:
    """Fetch latest spot FX rates (LCU per USD) for all relevant currencies.
    
    Returns dict mapping currency code -> spot rate (LCU per USD).
    For USD, returns 1.0.
    """
    out: dict[str, float] = {"USD": 1.0}
    for currency in sorted(currencies):
        ccy = normalize_currency_code(currency)
        if not ccy or ccy == "USD":
            continue

        spot_rate: float | None = None
        # Try both conventions: CCYUSD=X (gives USD per CCY) and USDCCY=X (gives CCY per USD)
        for pair, invert in ((f"{ccy}USD=X", True), (f"USD{ccy}=X", False)):
            try:
                tk = yf.Ticker(pair)
                hist = tk.history(period="5d", interval="1d", auto_adjust=True, actions=False)
                if not hist.empty and "Close" in hist.columns:
                    s = hist["Close"].dropna()
                    if not s.empty:
                        latest = float(s.iloc[-1])
                        if invert:
                            latest = 1.0 / latest
                        spot_rate = latest
                        break
            except Exception:
                continue
        if spot_rate is not None:
            out[ccy] = spot_rate
    return out


def build_timeframe_rows(
    etf_csv: str,
    weo_csv: str,
    metadata_csv: str | None = "data/outputs/etf_ticker_metadata.csv",
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = pd.read_csv(etf_csv)
    raw["Date"] = pd.to_datetime(raw["Date"], errors="coerce")
    raw = raw.dropna(subset=["Date"]).sort_values("Date")

    ticker_to_country = build_ticker_country_map()
    ticker_to_currency = load_ticker_currency_map(metadata_csv)
    price_columns = choose_price_columns(raw.columns.tolist())
    (
        gdp_real_growth_map,
        gdp_nominal_lcu_growth_map,
        gdp_nominal_usd_growth_map,
        country_lcu_vs_usd_weo_map,
        gdp_current_usd_map,
        gdp_current_lcu_map,
        inflation_cpi_map,
    ) = load_gdp_growth_maps(weo_csv)
    if raw.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    min_year = int(raw["Date"].dt.year.min())
    max_year = int(raw["Date"].dt.year.max())
    currencies = {
        normalize_currency_code(ticker_to_currency.get(ticker, ""))
        for ticker in price_columns
        if ticker_to_currency.get(ticker, "")
    }
    
    # Pre-fetch all needed data
    quote_fx_map = compute_annual_fx_quote_to_usd_returns(currencies, min_year, max_year)
    quote_fx_jan1_map = compute_annual_jan1_fx_returns(currencies, min_year, max_year)
    # Country LCU Jan 1 FX returns (for FX Jan 1 CAGR column - matches FX CAGR methodology)
    country_lcu_currencies = set(COUNTRY_TO_LCU.values())
    country_lcu_fx_jan1_map = compute_annual_jan1_fx_returns(country_lcu_currencies, min_year, max_year)
    daily_fx_series = fetch_daily_fx_series(currencies)

    timeframe_rows: list[dict[str, object]] = []
    annual_rows: list[dict[str, object]] = []
    cagr_rows: list[dict[str, object]] = []

    for ticker, price_col in price_columns.items():
        country_name = ticker_to_country.get(ticker)
        if not country_name:
            continue
        country_code = COUNTRY_TO_ISO3.get(country_name)
        if not country_code:
            continue

        series = raw.set_index("Date")[price_col].dropna()
        if series.empty:
            continue
        
        # Identify the first date with a non-zero price (inception)
        valid_series = series[series > 0]
        if valid_series.empty:
            continue
        inception_date = valid_series.index.min()
        inception_year = int(inception_date.year)

        end_pt = PricePoint(series.index[-1], float(series.iloc[-1]))
        end_date = end_pt.date
        completed_year = pd.Timestamp.today().year - 1
        etf_year_max = int(series.index.year.max())
        gdp_years = sorted(
            {
                year
                for (cc, year), _ in gdp_real_growth_map.items()
                if cc == country_code
            }
        )
        gdp_year_max = max(gdp_years) if gdp_years else completed_year
        cagr_end_year = min(completed_year, etf_year_max, gdp_year_max)
        
        ccy_norm = normalize_currency_code(ticker_to_currency.get(ticker, ""))
        fx_s = daily_fx_series.get(ccy_norm)

        # Relative + MAX windows
        for label, spec in TIMEFRAME_SPECS:
            if spec == "max":
                start_pt = max_start_point(series)
            elif spec == "ytd":
                jan1 = pd.Timestamp(year=end_date.year, month=1, day=1)
                start_pt = last_on_or_before(series, jan1)
            else:
                target = end_date - spec
                start_pt = last_on_or_before(series, target)

            if start_pt is None:
                continue

            etf_return = pct_return(start_pt.value, end_pt.value)
            
            # Calculate USD return for timeframe
            etf_return_usd = None
            if ccy_norm == "USD":
                etf_return_usd = etf_return
            elif fx_s is not None:
                fx_start = last_on_or_before(fx_s, start_pt.date)
                fx_end = last_on_or_before(fx_s, end_pt.date)
                if fx_start and fx_end:
                    fx_ret = pct_return(fx_start.value, fx_end.value)
                    etf_return_usd = (((1.0 + etf_return / 100.0) * (1.0 + fx_ret / 100.0)) - 1.0) * 100.0

            timeframe_rows.append(
                {
                    "country_name": country_name,
                    "country_code": country_code,
                    "ticker": ticker,
                    "timeframe": label,
                    "start_date": start_pt.date.date(),
                    "end_date": end_pt.date.date(),
                    "etf_return_pct": etf_return,
                    "etf_return_usd_pct": etf_return_usd,
                    "etf_currency": ticker_to_currency.get(ticker, ""),
                }
            )

        # Annual source rows
        if gdp_years:
            for y in gdp_years:
                start_pt = None
                end_year_pt = None
                etf_return = None
                etf_return_usd = None
                
                if y >= inception_year:
                    start_pt = first_in_year(series, y)
                    end_year_pt = last_in_year(series, y)
                    if start_pt is not None and end_year_pt is not None and start_pt.value > 0:
                        etf_return = pct_return(start_pt.value, end_year_pt.value)
                
                gdp_real_same = gdp_real_growth_map.get((country_code, y))
                gdp_nominal_lcu_same = gdp_nominal_lcu_growth_map.get((country_code, y))
                gdp_nominal_usd_same = gdp_nominal_usd_growth_map.get((country_code, y))
                gdp_current_usd_val = gdp_current_usd_map.get((country_code, y))
                quote_ccy_vs_usd = quote_fx_map.get((ccy_norm, y))
                quote_ccy_vs_usd_jan1 = quote_fx_jan1_map.get((ccy_norm, y))
                
                # Inflation data for Annual sheet
                inf_val = inflation_cpi_map.get((country_code, y))
                usa_inf_val = inflation_cpi_map.get(("USA", y))
                inf_diff = (inf_val - usa_inf_val) if (inf_val is not None and usa_inf_val is not None) else None

                if etf_return is not None:
                    if ccy_norm == "USD":
                        etf_return_usd = etf_return
                    elif quote_ccy_vs_usd is not None:
                        etf_return_usd = (
                            ((1.0 + (etf_return / 100.0)) * (1.0 + (quote_ccy_vs_usd / 100.0)))
                            - 1.0
                        ) * 100.0

                country_lcu_vs_usd_weo = country_lcu_vs_usd_weo_map.get((country_code, y))
                etf_usd_minus_country_fx = None
                if (
                    etf_return_usd is not None
                    and country_lcu_vs_usd_weo is not None
                    and (1.0 + (country_lcu_vs_usd_weo / 100.0)) != 0.0
                ):
                    etf_usd_minus_country_fx = (
                        ((1.0 + (etf_return_usd / 100.0))
                         / (1.0 + (country_lcu_vs_usd_weo / 100.0)))
                        - 1.0
                    ) * 100.0
                annual_rows.append(
                    {
                        "country_name": country_name,
                        "country_code": country_code,
                        "ticker": ticker,
                        "year": y,
                        "etf_return_pct": etf_return,
                        "etf_return_quote_pct": etf_return,
                        "quote_ccy_vs_usd_pct": quote_ccy_vs_usd,
                        "quote_ccy_vs_usd_jan1_pct": quote_ccy_vs_usd_jan1,
                        "etf_return_usd_pct": etf_return_usd,
                        "country_lcu_vs_usd_weo_pct": country_lcu_vs_usd_weo,
                        "etf_usd_minus_country_fx_pct": etf_usd_minus_country_fx,
                        "gdp_real_growth_pct": gdp_real_same,
                        "gdp_nominal_lcu_growth_pct": gdp_nominal_lcu_same,
                        "gdp_nominal_usd_growth_pct": gdp_nominal_usd_same,
                        "gdp_current_usd": gdp_current_usd_val,
                        "gdp_nominal_usd_minus_etf_growth_pct": (
                            gdp_nominal_usd_same - etf_return_usd
                            if (etf_return_usd is not None and gdp_nominal_usd_same is not None)
                            else None
                        ),
                        "inflation_cpi_pct": inf_val,
                        "usa_inflation_cpi_pct": usa_inf_val,
                        "inflation_cpi_diff": inf_diff,
                        "etf_currency": ticker_to_currency.get(ticker, ""),
                    }
                )

        # CAGR rows
        for years in CAGR_HORIZONS:
            start_year = cagr_end_year - years + 1
            if start_year < inception_year:
                continue
            
            start_pt = first_in_year(series, start_year)
            end_year_pt = last_in_year(series, cagr_end_year)
            if start_pt is None or end_year_pt is None or start_pt.value <= 0:
                continue
            
            total_ret_local = pct_return(start_pt.value, end_year_pt.value)
            
            if ccy_norm == "USD":
                total_ret_usd = total_ret_local
            else:
                fx_compounded = 1.0
                valid_fx = True
                for y in range(start_year, cagr_end_year + 1):
                    ann_fx = quote_fx_map.get((ccy_norm, y))
                    if ann_fx is None:
                        valid_fx = False
                        break
                    fx_compounded *= (1.0 + (ann_fx / 100.0))
                
                if valid_fx:
                    total_ret_usd = ((1.0 + (total_ret_local / 100.0)) * fx_compounded - 1.0) * 100.0
                else:
                    total_ret_usd = None
            
            if total_ret_usd is not None:
                etf_cagr = cagr_from_total_return(total_ret_usd, years)
            else:
                etf_cagr = None

            gdp_real_cagr = gdp_cagr(gdp_real_growth_map, country_code, cagr_end_year, years)
            gdp_nominal_lcu_cagr = gdp_cagr(gdp_nominal_lcu_growth_map, country_code, cagr_end_year, years)
            gdp_nominal_usd_cagr = gdp_cagr(gdp_nominal_usd_growth_map, country_code, cagr_end_year, years)
            
            # New metrics (FIXED: Ensure these are calculated consistently for all rows)
            fx_cagr_val = fx_level_cagr(gdp_current_lcu_map, gdp_current_usd_map, country_code, cagr_end_year, years)
            inf_cagr_local, inf_cagr_usa, inf_diff_cagr_val = calculate_inflation_metrics(inflation_cpi_map, country_code, cagr_end_year, years)

            # Jan 1st FX CAGR (using Country LCU vs USD, consistent with FX CAGR methodology)
            fx_jan1_cagr_val = None
            country_lcu = COUNTRY_TO_LCU.get(country_name)
            if country_lcu and country_lcu != "USD":
                jan1_fx_compounded = 1.0
                valid_jan1_fx = True
                for y in range(start_year, cagr_end_year + 1):
                    ann_fx_jan1 = country_lcu_fx_jan1_map.get((country_lcu, y))
                    if ann_fx_jan1 is None:
                        valid_jan1_fx = False
                        break
                    jan1_fx_compounded *= (1.0 + (ann_fx_jan1 / 100.0))

                if valid_jan1_fx:
                    fx_jan1_cagr_val = cagr_from_total_return((jan1_fx_compounded - 1.0) * 100.0, years)
            else:
                fx_jan1_cagr_val = 0.0

            cagr_rows.append(
                {
                    "country_name": country_name,
                    "country_code": country_code,
                    "ticker": ticker,
                    "as_of_date": end_year_pt.date.date(),
                    "horizon": f"{years}Y",
                    "etf_cagr_pct": etf_cagr,
                    "gdp_real_cagr_pct": gdp_real_cagr,
                    "gdp_nominal_lcu_cagr_pct": gdp_nominal_lcu_cagr,
                    "gdp_nominal_usd_cagr_pct": gdp_nominal_usd_cagr,
                    "fx_cagr_pct": fx_cagr_val,
                    "fx_jan1_cagr_pct": fx_jan1_cagr_val,
                    "inflation_local_cagr_pct": inf_cagr_local,
                    "inflation_usa_cagr_pct": inf_cagr_usa,
                    "inflation_diff_cagr_pct": inf_diff_cagr_val,
                    "gdp_nominal_usd_minus_etf_cagr_pct": (
                        gdp_nominal_usd_cagr - etf_cagr
                        if (gdp_nominal_usd_cagr is not None and etf_cagr is not None)
                        else None
                    ),
                    "etf_currency": ticker_to_currency.get(ticker, ""),
                }
            )

    timeframe_df = pd.DataFrame(timeframe_rows)
    if not timeframe_df.empty:
        timeframe_df["timeframe"] = pd.Categorical(
            timeframe_df["timeframe"], categories=TIMEFRAME_ORDER, ordered=True
        )
        timeframe_df = timeframe_df.sort_values(
            ["country_name", "ticker", "timeframe"]
        ).reset_index(drop=True)
        timeframe_df["timeframe"] = timeframe_df["timeframe"].astype(str)
        timeframe_df["lookup_key"] = (
            timeframe_df["country_name"] + "|" + timeframe_df["ticker"] + "|" + timeframe_df["timeframe"]
        )
        timeframe_df = timeframe_df[
            [
                "country_name",
                "country_code",
                "ticker",
                "timeframe",
                "start_date",
                "end_date",
                "etf_return_pct",
                "etf_return_usd_pct",
                "lookup_key",
                "etf_currency",
            ]
        ]

    annual_df = pd.DataFrame(annual_rows)
    if not annual_df.empty:
        annual_df = annual_df.sort_values(
            ["country_name", "ticker", "year"]
        ).reset_index(drop=True)
        annual_df["lookup_key"] = (
            annual_df["country_name"] + "|" + annual_df["ticker"] + "|" + annual_df["year"].astype(str)
        )
        annual_df["country_year_key"] = (
            annual_df["country_name"] + "|" + annual_df["year"].astype(str)
        )
        annual_df = annual_df[
            [
                "country_name",
                "country_code",
                "ticker",
                "year",
                "etf_return_pct",
                "gdp_real_growth_pct",
                "gdp_nominal_lcu_growth_pct",
                "gdp_nominal_usd_growth_pct",
                "gdp_nominal_usd_minus_etf_growth_pct",
                "lookup_key",
                "country_year_key",
                "etf_currency",
                "etf_return_quote_pct",
                "quote_ccy_vs_usd_pct",
                "quote_ccy_vs_usd_jan1_pct",
                "etf_return_usd_pct",
                "country_lcu_vs_usd_weo_pct",
                "etf_usd_minus_country_fx_pct",
                "inflation_cpi_pct",
                "usa_inflation_cpi_pct",
                "inflation_cpi_diff",
                "gdp_current_usd",
            ]
        ]

    cagr_df = pd.DataFrame(cagr_rows)
    gdp_only_rows: list[dict[str, object]] = []
    completed_year = pd.Timestamp.today().year - 1
    countries_in_scope = sorted(timeframe_df["country_name"].dropna().unique().tolist()) if not timeframe_df.empty else []
    for country_name in countries_in_scope:
        country_code = COUNTRY_TO_ISO3.get(country_name)
        if not country_code:
            continue
        gdp_years = sorted(
            {year for (cc, year), _ in gdp_real_growth_map.items() if cc == country_code}
        )
        if not gdp_years:
            continue
        cagr_end_year = min(completed_year, max(gdp_years))
        for years in CAGR_HORIZONS:
            gdp_real_cagr = gdp_cagr(gdp_real_growth_map, country_code, cagr_end_year, years)
            gdp_nominal_lcu_cagr = gdp_cagr(gdp_nominal_lcu_growth_map, country_code, cagr_end_year, years)
            gdp_nominal_usd_cagr = gdp_cagr(gdp_nominal_usd_growth_map, country_code, cagr_end_year, years)
            
            # FIXED: Correct calculation for GDP-only rows
            fx_cagr_val = fx_level_cagr(gdp_current_lcu_map, gdp_current_usd_map, country_code, cagr_end_year, years)
            inf_cagr_local, inf_cagr_usa, inf_diff_cagr_val = calculate_inflation_metrics(inflation_cpi_map, country_code, cagr_end_year, years)

            # Jan 1st FX CAGR for GDP-only rows (using Country LCU vs USD, consistent with FX CAGR)
            country_lcu = COUNTRY_TO_LCU.get(country_name)
            fx_jan1_cagr_val = None
            if country_lcu and country_lcu != "USD":
                jan1_fx_compounded = 1.0
                valid_jan1_fx = True
                start_year_gdp = cagr_end_year - years + 1
                for y in range(start_year_gdp, cagr_end_year + 1):
                    ann_fx_jan1 = country_lcu_fx_jan1_map.get((country_lcu, y))
                    if ann_fx_jan1 is None:
                        valid_jan1_fx = False
                        break
                    jan1_fx_compounded *= (1.0 + (ann_fx_jan1 / 100.0))
                if valid_jan1_fx:
                    fx_jan1_cagr_val = cagr_from_total_return((jan1_fx_compounded - 1.0) * 100.0, years)
            else:
                fx_jan1_cagr_val = 0.0
            
            if (
                gdp_real_cagr is None
                and gdp_nominal_lcu_cagr is None
                and gdp_nominal_usd_cagr is None
            ):
                continue
            gdp_only_rows.append(
                {
                    "country_name": country_name,
                    "country_code": country_code,
                    "ticker": "",
                    "as_of_date": pd.Timestamp(year=cagr_end_year, month=12, day=31).date(),
                    "horizon": f"{years}Y",
                    "etf_cagr_pct": None,
                    "gdp_real_cagr_pct": gdp_real_cagr,
                    "gdp_nominal_lcu_cagr_pct": gdp_nominal_lcu_cagr,
                    "gdp_nominal_usd_cagr_pct": gdp_nominal_usd_cagr,
                    "fx_cagr_pct": fx_cagr_val,
                    "fx_jan1_cagr_pct": fx_jan1_cagr_val,
                    "inflation_local_cagr_pct": inf_cagr_local,
                    "inflation_usa_cagr_pct": inf_cagr_usa,
                    "inflation_diff_cagr_pct": inf_diff_cagr_val,
                    "gdp_nominal_usd_minus_etf_cagr_pct": None,
                    "etf_currency": "",
                }
            )
    if gdp_only_rows:
        cagr_df = pd.concat([cagr_df, pd.DataFrame(gdp_only_rows)], ignore_index=True)

    if not cagr_df.empty:
        horizon_order = pd.Categorical(
            cagr_df["horizon"],
            categories=[f"{y}Y" for y in CAGR_HORIZONS],
            ordered=True,
        )
        cagr_df = (
            cagr_df.assign(_h=horizon_order)
            .sort_values(["country_name", "ticker", "_h"])
            .drop(columns=["_h"])
            .reset_index(drop=True)
        )
        cagr_df["lookup_key"] = (
            cagr_df["country_name"] + "|" + cagr_df["ticker"] + "|" + cagr_df["horizon"]
        )
        cagr_df["country_horizon_key"] = (
            cagr_df["country_name"] + "|" + cagr_df["horizon"]
        )
        cagr_df = cagr_df[
            [
                "country_name",
                "country_code",
                "ticker",
                "as_of_date",
                "horizon",
                "etf_cagr_pct",
                "gdp_real_cagr_pct",
                "gdp_nominal_lcu_cagr_pct",
                "gdp_nominal_usd_cagr_pct",
                "fx_cagr_pct",
                "fx_jan1_cagr_pct",
                "inflation_local_cagr_pct",
                "inflation_usa_cagr_pct",
                "inflation_diff_cagr_pct",
                "gdp_nominal_usd_minus_etf_cagr_pct",
                "lookup_key",
                "country_horizon_key",
                "etf_currency",
            ]
        ]

    return timeframe_df, annual_df, cagr_df


def write_comparing_countries_sheet(
    ws,
    cagr_df: pd.DataFrame,
    annual_df: pd.DataFrame,
    timeframe_df: pd.DataFrame,
    impact_df: pd.DataFrame,
    reer_df: pd.DataFrame,
    lists_df: pd.DataFrame,
    ticker_attrs_df: pd.DataFrame,
    country_summary_df: pd.DataFrame,
    spot_fx_rates: dict[str, float],
    gdp_current_usd_map: dict[tuple[str, int], float],
) -> None:
    """Write the 'Comparing Countries' sheet (country screener)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation

    # Helper to get column letter by dataframe column name
    def get_col(df, col_name):
        try:
            idx = df.columns.get_loc(col_name) + 1
            return get_column_letter(idx)
        except (ValueError, KeyError):
            return "A"

    c = {col: f"CAGR!${get_col(cagr_df, col)}:${get_col(cagr_df, col)}" for col in cagr_df.columns}
    a = {col: f"Annual!${get_col(annual_df, col)}:${get_col(annual_df, col)}" for col in annual_df.columns}

    # Build country-level summary using 5Y horizon as default
    horizon = "5Y"
    countries = sorted(cagr_df["country_name"].dropna().unique().tolist())
    
    # Header row
    ws["A1"] = "Country Disconnect Dashboard (As of " + pd.Timestamp.today().strftime("%b %d, %Y") + ")"
    ws["A1"].font = Font(bold=True, size=14)
    
    # Controls row
    ws["A2"] = "Horizon (choice between 1Y, 3Y, 5Y, 10Y)"
    ws["B2"] = "5Y"
    ws["D2"] = "Region Filter (using hidden reference columns)"
    ws["E2"] = "(asia, europe, etc.)"
    ws["G2"] = "Sort by (choice of GDP size)"
    ws["H2"] = "nGDP"
    
    # Section headers (row 4)
    ws["A4"] = "Country"
    ws["B4"] = "ETF Ticker"
    ws["C4"] = "nGDP (USD) CAGR (%)"
    ws["D4"] = "ETF Return (USD) CAGR (%)"
    ws["E4"] = "MSCI Index Retrun (USD) CAGR (%)"  # Placeholder - not available
    ws["F4"] = "Macro Gap (%)"
    ws["G4"] = "Projected Growth (2026-28) CAGR (%)"
    ws["H4"] = "Impact of $10 Oil Price change relative to GDP (%)"
    ws["I4"] = "REER Index"
    ws["J4"] = "REER vs 10Y"
    ws["K4"] = "Futures rate for USD/LCU 2 years out (%)"  # Placeholder - not available
    
    # GDP Metrics section
    ws["L4"] = "GDP Metrics"
    ws["L4"].font = Font(bold=True)
    ws["M4"] = "nGDP (USD) 2025"
    ws["N4"] = "nGDP (LCU) CAGR (%)"
    ws["O4"] = "rGDP (LCU) CAGR (%)"
    
    # Exchange Rate Metrics section
    ws["P4"] = "Exchange Rate Metrics"
    ws["P4"].font = Font(bold=True)
    ws["Q4"] = "FX CAGR %"
    ws["R4"] = "FX Jan 1st CAGR %"
    ws["S4"] = "USD/LCU spot rate as of \"today\""
    ws["T4"] = "Futures rate for USD/LCU (2 years out from \"today\")"  # Placeholder
    
    # Inflation Metrics section
    ws["U4"] = "Inflation Metrics"
    ws["U4"].font = Font(bold=True)
    ws["V4"] = "Inf. Diff CAGR %"
    
    # Differential section
    ws["W4"] = "Differential"
    ws["W4"].font = Font(bold=True)
    ws["X4"] = "Currency Gap %"
    
    # Country Info section
    ws["Y4"] = "Country Info"
    ws["Y4"].font = Font(bold=True)
    ws["Z4"] = "Region"
    ws["AA4"] = "Local Currency Unit (LCU)"
    
    # ETF Info section
    ws["AB4"] = "ETF Info"
    ws["AB4"].font = Font(bold=True)
    ws["AC4"] = "Exchange name for ETF ticker"
    ws["AD4"] = "ETF ticker currency"
    
    # Apply bold to all header cells
    for col_ref in ["A2", "A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4", "J4", "K4",
                    "M4", "N4", "O4", "Q4", "R4", "S4", "T4", "V4", "X4", "Z4", "AA4", "AC4", "AD4"]:
        ws[col_ref].font = Font(bold=True)
    
    # Add horizon dropdown
    horizon_dv = DataValidation(type="list", formula1='"1Y,3Y,5Y,10Y"', allow_blank=False)
    ws.add_data_validation(horizon_dv)
    horizon_dv.add("B2")
    
    # Write data rows
    country_start_row = 5
    country_count = len(countries)
    ticker_attrs_end_row = len(ticker_attrs_df) + 1
    
    for idx, country in enumerate(countries):
        r = country_start_row + idx
        
        # Get default ticker for this country (USD preferred, then largest AUM)
        country_tickers = timeframe_df[timeframe_df["country_name"] == country]["ticker"].unique().tolist()
        default_ticker = country_summary_df[country_summary_df["country_name"] == country]["ticker_used"].iloc[0] if not country_summary_df[country_summary_df["country_name"] == country].empty else (country_tickers[0] if country_tickers else "")
        
        ws[f"A{r}"] = country
        ws[f"B{r}"] = default_ticker
        
        # Add ticker dropdown
        country_ticker_options_df = lists_df[lists_df["country_name"] == country][["ticker_display"]].drop_duplicates()
        if not country_ticker_options_df.empty:
            ticker_list_range = f"Lists!$H$2:$H${len(lists_df) + 1}"
            ticker_formula = f'=OFFSET(Lists!$H$2,IFERROR(MATCH($A{r},Lists!$F$2:$F${len(lists_df)+1},0)-1,0),0,COUNTIF(Lists!$F$2:$F${len(lists_df)+1},$A{r}),1)'
            ticker_dv = DataValidation(type="list", formula1=ticker_formula, allow_blank=False)
            ws.add_data_validation(ticker_dv)
            ticker_dv.add(f"B{r}")
        
        # GDP CAGR (USD) - Column C
        ws[f"C{r}"] = f'=IFERROR(1*INDEX({c["gdp_nominal_usd_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # ETF CAGR (USD) - Column D
        ws[f"D{r}"] = f'=IFERROR(1*INDEX({c["etf_cagr_pct"]}, MATCH($A{r}&"|"&$B{r}&"|{horizon}", {c["lookup_key"]}, 0)), NA())'
        
        # MSCI Index Return - Column E (placeholder - not available)
        ws[f"E{r}"] = "NA()"
        
        # Macro Gap - Column F
        ws[f"F{r}"] = f"=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),C{r}-D{r},NA())"
        
        # Projected Growth 2026-28 - Column G
        idx26 = f'INDEX({a["gdp_nominal_usd_growth_pct"]}, MATCH($A{r}&"|2026", {a["country_year_key"]}, 0))'
        idx27 = f'INDEX({a["gdp_nominal_usd_growth_pct"]}, MATCH($A{r}&"|2027", {a["country_year_key"]}, 0))'
        idx28 = f'INDEX({a["gdp_nominal_usd_growth_pct"]}, MATCH($A{r}&"|2028", {a["country_year_key"]}, 0))'
        cagr_form = f'=IFERROR((( (1+{idx26}/100)*(1+{idx27}/100)*(1+{idx28}/100) )^(1/3)-1)*100, NA())'
        ws[f"G{r}"] = cagr_form
        
        # Oil Impact - Column H
        ws[f"H{r}"] = f'=IFERROR(INDEX(Crude_Oil_Impact!$H:$H, MATCH($A{r}, Crude_Oil_Impact!$A:$A, 0)), NA())'
        
        # REER Index - Column I
        ws[f"I{r}"] = f'=IFERROR(INDEX(REER_Data!$C:$C, MATCH($A{r}, REER_Data!$A:$A, 0)), NA())'
        
        # REER vs 10Y - Column J (divide by 100 for % formatting)
        ws[f"J{r}"] = f'=IFERROR(INDEX(REER_Data!$E:$E, MATCH($A{r}, REER_Data!$A:$A, 0)) / 100, NA())'
        
        # Futures rate - Column K (placeholder)
        ws[f"K{r}"] = "NA()"
        
        # nGDP (USD) 2025 - Column M (from WEO levels)
        country_code = COUNTRY_TO_ISO3.get(country)
        gdp_2025_val = gdp_current_usd_map.get((country_code, 2025)) if country_code else None
        if gdp_2025_val is not None:
            ws[f"M{r}"] = gdp_2025_val
        else:
            ws[f"M{r}"] = f'=IFERROR(INDEX({a["gdp_current_usd"]}, MATCH($A{r}&"|2025", {a["country_year_key"]}, 0)), NA())'
        
        # FX CAGR - Column Q
        ws[f"Q{r}"] = f'=IFERROR(1*INDEX({c["fx_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # FX Jan 1st CAGR - Column R
        ws[f"R{r}"] = f'=IFERROR(1*INDEX({c["fx_jan1_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # Spot FX rate - Column S
        country_ticker_data = timeframe_df[timeframe_df["country_name"] == country][["ticker", "etf_currency"]].drop_duplicates()
        if not country_ticker_data.empty:
            currency = country_ticker_data["etf_currency"].iloc[0]
            spot_rate = spot_fx_rates.get(currency)
            if spot_rate is not None:
                ws[f"S{r}"] = spot_rate
        
        # Futures rate 2y out - Column T (placeholder)
        ws[f"T{r}"] = "NA()"
        
        # Inf. Diff CAGR - Column V
        ws[f"V{r}"] = f'=IFERROR(1*INDEX({c["inflation_diff_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # Currency Gap - Column X
        ws[f"X{r}"] = f"=IF(AND(ISNUMBER(Q{r}),ISNUMBER(V{r})),Q{r}+V{r},NA())"
        
        # Region - Column Z
        region = COUNTRY_TO_REGION.get(country, "Other")
        ws[f"Z{r}"] = region
        
        # LCU - Column AA
        lcu = COUNTRY_TO_LCU.get(country, "")
        ws[f"AA{r}"] = lcu
        
        # Exchange - Column AC
        ws[f"AC{r}"] = f'=IFERROR(INDEX(Lists!$J$2:$J${ticker_attrs_end_row}, MATCH($B{r}, Lists!$I$2:$I${ticker_attrs_end_row}, 0)),"")'
        
        # Currency - Column AD
        ws[f"AD{r}"] = f'=IFERROR(INDEX(Lists!$K$2:$K${ticker_attrs_end_row}, MATCH($B{r}, Lists!$I$2:$I${ticker_attrs_end_row}, 0)),"")'
        
        # nGDP (LCU) CAGR - Column N
        ws[f"N{r}"] = f'=IFERROR(1*INDEX({c["gdp_nominal_lcu_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # rGDP (LCU) CAGR - Column O
        ws[f"O{r}"] = f'=IFERROR(1*INDEX({c["gdp_real_cagr_pct"]}, MATCH($A{r}&"|{horizon}", {c["country_horizon_key"]}, 0)), NA())'
        
        # Apply number formats
        for col in ["C", "D", "F", "G", "H", "J", "N", "O", "Q", "R", "V", "X"]:
            ws[f"{col}{r}"].number_format = "0.00"
        ws[f"I{r}"].number_format = "0.00"
        ws[f"S{r}"].number_format = "0.0000"
    
    # Apply conditional formatting
    green_light = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    green_medium = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    green_saturated = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    red_light = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    red_saturated = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    country_end_row = country_start_row + country_count - 1
    
    # Main metrics formatting
    format_ranges = [
        f"C{country_start_row}:D{country_end_row}",  # GDP/ETF CAGR
        f"F{country_start_row}:G{country_end_row}",  # Macro Gap, Proj 3Y
        f"J{country_start_row}:J{country_end_row}",  # REER vs 10Y
        f"N{country_start_row}:O{country_end_row}",  # GDP LCU/Real CAGR
        f"Q{country_start_row}:R{country_end_row}",  # FX CAGR
        f"V{country_start_row}:V{country_end_row}",  # Inf Diff
        f"X{country_start_row}:X{country_end_row}",  # Currency Gap
    ]
    for rng in format_ranges:
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["10"], fill=green_saturated))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=green_medium))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_light))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=red_light))
    
    # Oil Impact formatting (lower is better)
    oil_range = f"H{country_start_row}:H{country_end_row}"
    ws.conditional_formatting.add(oil_range, CellIsRule(operator="lessThanOrEqual", formula=["1"], fill=green_saturated))
    ws.conditional_formatting.add(oil_range, CellIsRule(operator="lessThanOrEqual", formula=["3"], fill=green_light))
    ws.conditional_formatting.add(oil_range, CellIsRule(operator="greaterThan", formula=["3"], fill=red_light))
    ws.conditional_formatting.add(oil_range, CellIsRule(operator="greaterThan", formula=["6"], fill=red_saturated))
    
    # REER Index formatting
    reer_range = f"I{country_start_row}:I{country_end_row}"
    ws.conditional_formatting.add(reer_range, CellIsRule(operator="greaterThanOrEqual", formula=["110"], fill=green_saturated))
    ws.conditional_formatting.add(reer_range, CellIsRule(operator="between", formula=["100", "110"], fill=green_light))
    ws.conditional_formatting.add(reer_range, CellIsRule(operator="between", formula=["90", "100"], fill=red_light))
    ws.conditional_formatting.add(reer_range, CellIsRule(operator="lessThan", formula=["90"], fill=red_saturated))
    
    # Apply yellow fill to BOE Price Impact for UN fallback countries
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if not impact_df.empty and "data_source" in impact_df.columns:
        un_countries = set(impact_df[impact_df["data_source"] == "UN"]["country_name"].tolist())
        for r in range(country_start_row, country_end_row + 1):
            country = ws[f"A{r}"].value
            if country in un_countries:
                ws.cell(row=r, column=8).fill = yellow_fill  # Column H
    
    # Set column widths
    ws.column_dimensions["A"].width = 18  # Country
    ws.column_dimensions["B"].width = 12  # Ticker
    ws.column_dimensions["C"].width = 15  # GDP CAGR
    ws.column_dimensions["D"].width = 15  # ETF CAGR
    ws.column_dimensions["E"].width = 15  # MSCI (placeholder)
    ws.column_dimensions["F"].width = 12  # Macro Gap
    ws.column_dimensions["G"].width = 15  # Proj 3Y
    ws.column_dimensions["H"].width = 15  # Oil Impact
    ws.column_dimensions["I"].width = 12  # REER Index
    ws.column_dimensions["J"].width = 12  # REER vs 10Y
    ws.column_dimensions["K"].width = 15  # Futures (placeholder)
    ws.column_dimensions["M"].width = 15  # nGDP 2025
    ws.column_dimensions["N"].width = 15  # nGDP LCU CAGR
    ws.column_dimensions["O"].width = 15  # rGDP LCU CAGR
    ws.column_dimensions["Q"].width = 12  # FX CAGR
    ws.column_dimensions["R"].width = 15  # FX Jan 1st
    ws.column_dimensions["S"].width = 15  # Spot FX
    ws.column_dimensions["T"].width = 15  # Futures 2y
    ws.column_dimensions["V"].width = 15  # Inf Diff
    ws.column_dimensions["X"].width = 15  # Currency Gap
    ws.column_dimensions["Z"].width = 12  # Region
    ws.column_dimensions["AA"].width = 12  # LCU
    ws.column_dimensions["AC"].width = 20  # Exchange
    ws.column_dimensions["AD"].width = 12  # Currency
    
    # Hide placeholder columns (E, K, T)
    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["K"].hidden = True
    ws.column_dimensions["T"].hidden = True
    
    # Set autofilter and freeze
    ws.auto_filter.ref = f"A4:AD{country_end_row}"
    ws.freeze_panes = "A5"


def write_country_focus_sheet(
    ws,
    timeframe_df: pd.DataFrame,
    annual_df: pd.DataFrame,
    cagr_df: pd.DataFrame,
    lists_df: pd.DataFrame,
    ticker_attrs_df: pd.DataFrame,
) -> None:
    """Write the 'Country Focus' sheet (detailed single-country view)."""
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation

    # Helper to get column letter by dataframe column name
    def get_col(df, col_name):
        try:
            idx = df.columns.get_loc(col_name) + 1
            return get_column_letter(idx)
        except (ValueError, KeyError):
            return "A"

    t = {col: f"ETF_Timeframes!${get_col(timeframe_df, col)}:${get_col(timeframe_df, col)}" for col in timeframe_df.columns}
    a = {col: f"Annual!${get_col(annual_df, col)}:${get_col(annual_df, col)}" for col in annual_df.columns}
    c = {col: f"CAGR!${get_col(cagr_df, col)}:${get_col(cagr_df, col)}" for col in cagr_df.columns}
    
    # Header
    ws["A1"] = "Country Focus Dashboard"
    ws["A1"].font = Font(bold=True, size=14)
    
    # Selectors
    ws["A2"] = "Country"
    ws["A3"] = "Ticker"
    ws["A4"] = "As-of Date"
    ws["A5"] = "Ticker Currency"
    ws["A6"] = "Ticker Exchange"
    
    # Country dropdown
    country_count = len(lists_df["country_name"].dropna().unique())
    country_dv = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${country_count + 1}", allow_blank=False)
    ws.add_data_validation(country_dv)
    country_dv.add("B2")
    countries_unique = timeframe_df["country_name"].dropna().unique()
    ws["B2"] = countries_unique[0] if len(countries_unique) > 0 else ""
    
    # Ticker dropdown (dynamic based on country)
    ticker_attrs_end_row = len(ticker_attrs_df) + 1
    country_ticker_end_row = len(lists_df) + 1
    ticker_formula = f'=OFFSET(Lists!$H$2,IFERROR(MATCH($B$2,Lists!$F$2:$F${country_ticker_end_row},0)-1,0),0,COUNTIF(Lists!$F$2:$F${country_ticker_end_row},$B$2),1)'
    ticker_dv = DataValidation(type="list", formula1=ticker_formula, allow_blank=False)
    ws.add_data_validation(ticker_dv)
    ticker_dv.add("B3")
    ws["B3"] = f'=IFERROR(INDEX(Lists!$D:$D, MATCH($B$2, Lists!$C:$C, 0)),"")'
    
    # As-of Date
    ws["B4"] = f'=IFERROR(INDEX({t["end_date"]}, MATCH($B$2&"|"&$B$3&"|MAX", {t["lookup_key"]}, 0)),"")'
    ws["B4"].number_format = "yyyy-mm-dd"
    
    # Currency and Exchange
    ws["B5"] = f'=IFERROR(INDEX(Lists!$K$2:$K${ticker_attrs_end_row}, MATCH($B$3, Lists!$I$2:$I${ticker_attrs_end_row}, 0)),"")'
    ws["B6"] = f'=IFERROR(INDEX(Lists!$J$2:$J${ticker_attrs_end_row}, MATCH($B$3, Lists!$I$2:$I${ticker_attrs_end_row}, 0)),"")'
    
    focus_currency_ref = "$B$5"
    
    # Section 1: ETF Cumulative Returns
    timeframe_title_row = 8
    timeframe_header_row = 9
    timeframe_start_row = 10
    
    ws[f"A{timeframe_title_row}"] = "ETF Cumulative returns"
    ws[f"A{timeframe_title_row}"].font = Font(bold=True)
    ws[f"A{timeframe_header_row}"] = "Timeframe"
    ws[f"B{timeframe_header_row}"] = "ETF Return (USD) %"
    ws[f"C{timeframe_header_row}"] = f'=IF({focus_currency_ref}<>"USD","ETF Return (" & {focus_currency_ref} & ") %","")'
    ws[f"D{timeframe_header_row}"] = "Start Date"
    
    for hdr_col in ["A", "B", "C", "D"]:
        ws[f"{hdr_col}{timeframe_header_row}"].font = Font(bold=True)
    
    for i, tf in enumerate(TIMEFRAME_ORDER):
        r = timeframe_start_row + i
        ws[f"A{r}"] = tf
        ws[f"B{r}"] = f'=IFERROR(INDEX({t["etf_return_usd_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {t["lookup_key"]}, 0)), "")'
        ws[f"C{r}"] = f'=IF({focus_currency_ref}<>"USD", IFERROR(INDEX({t["etf_return_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {t["lookup_key"]}, 0)), ""), "")'
        ws[f"D{r}"] = f'=IFERROR(INDEX({t["start_date"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {t["lookup_key"]}, 0)),"")'
        ws[f"B{r}"].number_format = "0.00"
        ws[f"C{r}"].number_format = "0.00"
        ws[f"D{r}"].number_format = "yyyy-mm-dd"
    
    # Section 2: Annual ETF vs GDP + FX Decomposition
    annual_title_row = timeframe_start_row + len(TIMEFRAME_ORDER) + 3
    annual_header_row = annual_title_row + 1
    annual_start_row = annual_header_row + 1
    
    ws[f"A{annual_title_row}"] = "Annual ETF vs GDP + FX Decomposition (Last 10 Years, %)"
    ws[f"A{annual_title_row}"].font = Font(bold=True)
    ws[f"A{annual_header_row}"] = "Year"
    ws[f"B{annual_header_row}"] = "Nominal GDP Growth (USD) %"
    ws[f"C{annual_header_row}"] = "ETF Return (USD) %"
    ws[f"D{annual_header_row}"] = "Macro Disconnect %"
    ws[f"E{annual_header_row}"] = ""  # GAP
    ws[f"F{annual_header_row}"] = "Real GDP Growth %"
    ws[f"G{annual_header_row}"] = "Nominal GDP Growth (LCU) %"
    ws[f"H{annual_header_row}"] = f'=IF({focus_currency_ref}<>"USD","ETF Return (" & {focus_currency_ref} & ") %","")'
    ws[f"I{annual_header_row}"] = f'=IF({focus_currency_ref}<>"USD","FX Change (vs USD) %","")'
    ws[f"J{annual_header_row}"] = "WEO LCU Return vs USD %"
    ws[f"K{annual_header_row}"] = f'=IF({focus_currency_ref}<>"USD","Jan 1st FX Change %","")'
    
    for hdr_col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws[f"{hdr_col}{annual_header_row}"].font = Font(bold=True)
    
    completed_year = pd.Timestamp.today().year - 1
    annual_years = sorted([y for y in annual_df["year"].dropna().unique().tolist() if int(y) <= completed_year])[-ANNUAL_WINDOW_YEARS:]
    
    for i, year in enumerate(annual_years):
        r = annual_start_row + i
        ws[f"A{r}"] = int(year)
        ws[f"B{r}"] = f'=IFERROR(1*INDEX({a["gdp_nominal_usd_growth_pct"]}, MATCH($B$2&"|"&$A{r}, {a["country_year_key"]}, 0)), NA())'
        ws[f"C{r}"] = f'=IFERROR(1*INDEX({a["etf_return_usd_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {a["lookup_key"]}, 0)), NA())'
        ws[f"D{r}"] = f"=IF(AND(ISNUMBER(B{r}),ISNUMBER(C{r}),C{r}<>0),B{r}-C{r},NA())"
        ws[f"E{r}"] = ""
        ws[f"F{r}"] = f'=IFERROR(1*INDEX({a["gdp_real_growth_pct"]}, MATCH($B$2&"|"&$A{r}, {a["country_year_key"]}, 0)), NA())'
        ws[f"G{r}"] = f'=IFERROR(1*INDEX({a["gdp_nominal_lcu_growth_pct"]}, MATCH($B$2&"|"&$A{r}, {a["country_year_key"]}, 0)), NA())'
        ws[f"H{r}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["etf_return_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {a["lookup_key"]}, 0)), NA()),"")'
        ws[f"I{r}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["quote_ccy_vs_usd_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {a["lookup_key"]}, 0)), NA()),"")'
        ws[f"J{r}"] = f'=IFERROR(1*INDEX({a["country_lcu_vs_usd_weo_pct"]}, MATCH($B$2&"|"&$A{r}, {a["country_year_key"]}, 0)), NA())'
        ws[f"K{r}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["quote_ccy_vs_usd_jan1_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {a["lookup_key"]}, 0)), NA()),"")'
    
    # Projection year row
    projection_years = sorted([y for y in annual_df["year"].dropna().unique().tolist() if int(y) > completed_year])
    annual_last_row = annual_header_row + len(annual_years)
    if projection_years:
        projection_year = int(projection_years[0])
        projection_row = annual_start_row + len(annual_years)
        ws[f"A{projection_row}"] = f"{projection_year} (Proj GDP / ETF YTD)"
        ws[f"B{projection_row}"] = f'=IFERROR(1*INDEX({a["gdp_nominal_usd_growth_pct"]}, MATCH($B$2&"|"&{projection_year}, {a["country_year_key"]}, 0)), NA())'
        proj_usd_ret_idx = f'INDEX({t["etf_return_usd_pct"]}, MATCH($B$2&"|"&$B$3&"|YTD", {t["lookup_key"]}, 0))'
        ws[f"C{projection_row}"] = f"=IFERROR(1*{proj_usd_ret_idx}, NA())"
        ws[f"D{projection_row}"] = f"=IF(AND(ISNUMBER(B{projection_row}),ISNUMBER(C{projection_row}),C{projection_row}<>0),B{projection_row}-C{projection_row},NA())"
        ws[f"E{projection_row}"] = ""
        ws[f"F{projection_row}"] = f'=IFERROR(1*INDEX({a["gdp_real_growth_pct"]}, MATCH($B$2&"|"&{projection_year}, {a["country_year_key"]}, 0)), NA())'
        ws[f"G{projection_row}"] = f'=IFERROR(1*INDEX({a["gdp_nominal_lcu_growth_pct"]}, MATCH($B$2&"|"&{projection_year}, {a["country_year_key"]}, 0)), NA())'
        ws[f"H{projection_row}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["etf_return_pct"]}, MATCH($B$2&"|"&$B$3&"|"&{projection_year}, {a["lookup_key"]}, 0)), NA()),"")'
        ws[f"I{projection_row}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["quote_ccy_vs_usd_pct"]}, MATCH($B$2&"|"&$B$3&"|"&{projection_year}, {a["lookup_key"]}, 0)), NA()),"")'
        ws[f"J{projection_row}"] = f'=IFERROR(1*INDEX({a["country_lcu_vs_usd_weo_pct"]}, MATCH($B$2&"|"&{projection_year}, {a["country_year_key"]}, 0)), NA())'
        ws[f"K{projection_row}"] = f'=IF({focus_currency_ref}<>"USD",IFERROR(1*INDEX({a["quote_ccy_vs_usd_jan1_pct"]}, MATCH($B$2&"|"&$B$3&"|"&{projection_year}, {a["lookup_key"]}, 0)), NA()),"")'
        annual_last_row = projection_row
        ws[f"A{annual_last_row + 1}"] = f"* {projection_year} row: GDP is IMF projection; ETF is YTD return (USD)."
    
    # Section 3: CAGR Comparison
    cagr_title_row = annual_last_row + 3
    cagr_header_row = cagr_title_row + 1
    cagr_start_row = cagr_header_row + 1
    
    ws[f"A{cagr_title_row}"] = "CAGR Comparison (%)"
    ws[f"A{cagr_title_row}"].font = Font(bold=True)
    ws[f"A{cagr_header_row}"] = "Horizon"
    ws[f"B{cagr_header_row}"] = "Nominal GDP CAGR % (USD)"
    ws[f"C{cagr_header_row}"] = "ETF CAGR % (USD)"
    ws[f"D{cagr_header_row}"] = ""  # GAP
    ws[f"E{cagr_header_row}"] = ""  # GAP
    ws[f"F{cagr_header_row}"] = "Real GDP CAGR %"
    ws[f"G{cagr_header_row}"] = "Nominal GDP CAGR % (LCU)"
    
    for hdr_col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws[f"{hdr_col}{cagr_header_row}"].font = Font(bold=True)
    
    for i, hz in enumerate(["3Y", "5Y", "10Y"]):
        r = cagr_start_row + i
        ws[f"A{r}"] = hz
        ws[f"B{r}"] = f'=IFERROR(1*INDEX({c["gdp_nominal_usd_cagr_pct"]}, MATCH($B$2&"|"&$A{r}, {c["country_horizon_key"]}, 0)), NA())'
        ws[f"C{r}"] = f'=IFERROR(1*INDEX({c["etf_cagr_pct"]}, MATCH($B$2&"|"&$B$3&"|"&$A{r}, {c["lookup_key"]}, 0)), NA())'
        ws[f"D{r}"] = ""
        ws[f"E{r}"] = ""
        ws[f"F{r}"] = f'=IFERROR(1*INDEX({c["gdp_real_cagr_pct"]}, MATCH($B$2&"|"&$A{r}, {c["country_horizon_key"]}, 0)), NA())'
        ws[f"G{r}"] = f'=IFERROR(1*INDEX({c["gdp_nominal_lcu_cagr_pct"]}, MATCH($B$2&"|"&$A{r}, {c["country_horizon_key"]}, 0)), NA())'
    
    # Apply number formats
    for row in range(annual_start_row, annual_last_row + 1):
        for col in ["B", "C", "D", "F", "G", "H", "I", "J"]:
            ws[f"{col}{row}"].number_format = "0.00"
    for row in range(cagr_start_row, cagr_start_row + 2):
        for col in ["B", "C", "F", "G"]:
            ws[f"{col}{row}"].number_format = "0.00"
    for row in range(timeframe_start_row, timeframe_start_row + len(TIMEFRAME_ORDER)):
        for col in ["B", "C"]:
            ws[f"{col}{row}"].number_format = "0.00"
    
    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 18


def write_dashboard_xlsx(
    timeframe_df: pd.DataFrame,
    annual_df: pd.DataFrame,
    cagr_df: pd.DataFrame,
    output_xlsx: str,
    metadata_csv: str | None = "data/outputs/etf_ticker_metadata.csv",
    reer_csv: str | None = "data/outputs/bis_reer_metrics.csv",
    impact_csv: str | None = "data/outputs/crude_oil_import_impact.csv",
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        ticker_fund_size_map = load_ticker_fund_size_map(metadata_csv)
        ticker_exchange_map = load_ticker_exchange_map(metadata_csv)
        
        # Load Crude Oil Impact data
        impact_df = pd.DataFrame()
        if impact_csv:
            try:
                impact_df = pd.read_csv(impact_csv)
                # Round and rename for display (BOE methodology with data source)
                display_impact = impact_df.copy()
                display_impact.columns = [
                    "Country", "ISO3",
                    "Crude Oil (Mbbl)", "Natural Gas (Mbbl BOE)", "Total BOE (Mbbl)",
                    "Value @ $10 Change (USD)", "Nominal GDP (USD)", "Impact (% of GDP)",
                    "Data Source"
                ]
                display_impact.to_excel(writer, sheet_name="Crude_Oil_Impact", index=False)
            except FileNotFoundError:
                pass

        default_map = (
            timeframe_df[timeframe_df["timeframe"] == "MAX"][
                ["country_name", "ticker", "etf_currency"]
            ]
            .assign(
                usd_rank=lambda d: (d["etf_currency"] == "USD").astype(int),
                fund_size_rank=lambda d: d["ticker"].map(ticker_fund_size_map).fillna(-1.0),
            )
            .sort_values(["country_name", "usd_rank", "fund_size_rank", "ticker"], ascending=[True, False, False, True])
            .groupby("country_name", as_index=False)
            .first()[["country_name", "ticker"]]
        )
        ticker_currency_map = (
            timeframe_df[["ticker", "etf_currency"]]
            .drop_duplicates(subset=["ticker"], keep="first")
            .set_index("ticker")["etf_currency"]
            .to_dict()
        )
        
        # Load ticker accumulating map and create display ticker function
        ticker_accumulating_map = load_ticker_accumulating_map(metadata_csv)
        
        def make_display_ticker(ticker):
            is_acc = ticker_accumulating_map.get(ticker, "yes")
            if is_acc == "no":
                return ticker + "*"
            return ticker

        country_summary_df = default_map.rename(columns={"ticker": "ticker_used"}).copy()
        # Add display ticker
        country_summary_df["ticker_used_display"] = country_summary_df["ticker_used"].apply(make_display_ticker)

        latest_gdp = (
            annual_df.dropna(subset=["gdp_current_usd"])
            .sort_values("year")
            .groupby("country_name")
            .last()[["gdp_current_usd"]]
        )
        country_summary_df = country_summary_df.merge(latest_gdp, on="country_name", how="left")

        country_summary_df["ticker_exchange"] = country_summary_df["ticker_used"].map(ticker_exchange_map).fillna("")
        country_summary_df["ticker_currency"] = country_summary_df["ticker_used"].map(ticker_currency_map).fillna("")
        country_summary_df["region"] = country_summary_df["country_name"].map(COUNTRY_TO_REGION).fillna("Other")
        country_summary_df = country_summary_df.sort_values("gdp_current_usd", ascending=False).reset_index(drop=True)
        country_summary_df_export = country_summary_df[["country_name", "ticker_used_display"]]

        country_ticker_options_df = (
            timeframe_df[["country_name", "ticker"]]
            .drop_duplicates(subset=["country_name", "ticker"], keep="first")
            .sort_values(["country_name", "ticker"], ascending=[True, True])
            .reset_index(drop=True)
        )

        ticker_attrs_df = (
            timeframe_df[["ticker"]]
            .drop_duplicates(subset=["ticker"], keep="first")
            .assign(
                ticker_display=lambda d: d["ticker"].apply(make_display_ticker),
                ticker_exchange=lambda d: d["ticker"].map(ticker_exchange_map).fillna(""),
                ticker_currency=lambda d: d["ticker"].map(ticker_currency_map).fillna(""),
            )
            .sort_values(["ticker"], ascending=[True])
            .reset_index(drop=True)
        )

        # Helper to get column letter by dataframe column name
        def get_col(df, col_name):
            try:
                idx = df.columns.get_loc(col_name) + 1
                return get_column_letter(idx)
            except (ValueError, KeyError):
                return "A" # Fallback

        # Column maps for formulas
        c = {col: f"CAGR!${get_col(cagr_df, col)}:${get_col(cagr_df, col)}" for col in cagr_df.columns}
        a = {col: f"Annual!${get_col(annual_df, col)}:${get_col(annual_df, col)}" for col in annual_df.columns}
        t = {col: f"ETF_Timeframes!${get_col(timeframe_df, col)}:${get_col(timeframe_df, col)}" for col in timeframe_df.columns}

        timeframe_df.to_excel(writer, sheet_name="ETF_Timeframes", index=False)
        annual_df.to_excel(writer, sheet_name="Annual", index=False)
        cagr_df.to_excel(writer, sheet_name="CAGR", index=False)
        
        reer_df = pd.DataFrame()
        if reer_csv:
            try:
                reer_df = pd.read_csv(reer_csv)
                reer_df.to_excel(writer, sheet_name="REER_Data", index=False)
            except FileNotFoundError:
                pass
        # Create default map with display ticker
        default_map_display = default_map.copy()
        default_map_display["ticker_display"] = default_map_display["ticker"].apply(make_display_ticker)
        
        lists_df = pd.DataFrame(
            {
                "country_name": pd.Series(sorted(timeframe_df["country_name"].dropna().unique().tolist())),
                "ticker": pd.Series(sorted(timeframe_df["ticker"].dropna().unique().tolist())),
                "country_for_default": pd.Series(default_map["country_name"].tolist()),
                "default_ticker": pd.Series(default_map_display["ticker_display"].tolist()),
            }
        )
        lists_df["region"] = lists_df["country_name"].map(COUNTRY_TO_REGION).fillna("Other")
        
        # Add ticker_display column for dropdown lookups
        lists_df["ticker_display"] = lists_df["ticker"].apply(make_display_ticker)

        lists_df.to_excel(writer, sheet_name="Lists", index=False)
        # Write country_ticker_options with display ticker
        country_ticker_options_df_with_display = country_ticker_options_df.copy()
        country_ticker_options_df_with_display["ticker_display"] = country_ticker_options_df_with_display["ticker"].apply(make_display_ticker)
        country_ticker_options_df_with_display.to_excel(writer, sheet_name="Lists", index=False, startcol=5)
        ticker_attrs_df.to_excel(writer, sheet_name="Lists", index=False, startcol=9)

        # Fetch spot FX rates for all currencies
        currencies = set(timeframe_df["etf_currency"].dropna().unique().tolist())
        spot_fx_rates = get_spot_fx_rates(currencies)

        # Build GDP current USD map for 2025 levels
        gdp_current_usd_map = {}
        if not annual_df.empty:
            for _, row in annual_df.iterrows():
                country_code = row.get("country_code")
                year = row.get("year")
                gdp_usd = row.get("gdp_current_usd")
                if country_code and year and pd.notna(gdp_usd):
                    gdp_current_usd_map[(str(country_code), int(year))] = float(gdp_usd)

        wb = writer.book
        ws_tf = wb["ETF_Timeframes"]
        ws_annual = wb["Annual"]
        ws_cagr = wb["CAGR"]
        ws_lists = wb["Lists"]

        if "Crude_Oil_Impact" in wb.sheetnames:
            ws_impact = wb["Crude_Oil_Impact"]
            ws_impact.freeze_panes = "A2"
            fit_columns_to_header(ws_impact, 1)
            # Apply number format to last column (Impact % of GDP)
            for cell in ws_impact[get_column_letter(ws_impact.max_column)]:
                if cell.row > 1:
                    cell.number_format = "0.000"

        ws_tf.auto_filter.ref = ws_tf.dimensions
        ws_tf.freeze_panes = "A2"
        ws_annual.auto_filter.ref = ws_annual.dimensions
        ws_annual.freeze_panes = "A2"
        ws_cagr.auto_filter.ref = ws_cagr.dimensions
        ws_cagr.freeze_panes = "A2"

        # Create new sheets: "Comparing Countries" and "Country Focus"
        ws_comparing = wb.create_sheet(title="Comparing Countries")
        ws_focus = wb.create_sheet(title="Country Focus")

        # Load REER data for passing to comparing countries sheet
        reer_df = pd.DataFrame()
        if reer_csv:
            try:
                reer_df = pd.read_csv(reer_csv)
            except FileNotFoundError:
                pass

        # Write the two new dashboard sheets
        write_comparing_countries_sheet(
            ws_comparing,
            cagr_df=cagr_df,
            annual_df=annual_df,
            timeframe_df=timeframe_df,
            impact_df=impact_df,
            reer_df=reer_df,
            lists_df=lists_df,
            ticker_attrs_df=ticker_attrs_df,
            country_summary_df=country_summary_df,
            spot_fx_rates=spot_fx_rates,
            gdp_current_usd_map=gdp_current_usd_map,
        )

        write_country_focus_sheet(
            ws_focus,
            timeframe_df=timeframe_df,
            annual_df=annual_df,
            cagr_df=cagr_df,
            lists_df=lists_df,
            ticker_attrs_df=ticker_attrs_df,
        )

        # Set active sheet to Comparing Countries
        wb.active = wb.sheetnames.index("Comparing Countries")
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Excel MVP dashboard for country-level ETF returns vs GDP growth.")
    parser.add_argument("--etf-csv", default="data/outputs/etf_prices.csv")
    parser.add_argument("--weo-csv", default="data/outputs/weo_gdp.csv")
    parser.add_argument("--metadata-csv", default="data/outputs/etf_ticker_metadata.csv")
    parser.add_argument("--reer-csv", default="data/outputs/bis_reer_metrics.csv")
    parser.add_argument("--crude-impact-csv", default="data/outputs/crude_oil_import_impact.csv")
    parser.add_argument("--output", default="data/outputs/etf_gdp_dashboard_mvp.xlsx")
    args = parser.parse_args()

    timeframe_df, annual_df, cagr_df = build_timeframe_rows(args.etf_csv, args.weo_csv, args.metadata_csv)
    if timeframe_df.empty or annual_df.empty or cagr_df.empty:
        raise RuntimeError("No rows produced. Check ETF and WEO input data.")
    write_dashboard_xlsx(
        timeframe_df, annual_df, cagr_df, args.output, 
        metadata_csv=args.metadata_csv, reer_csv=args.reer_csv, 
        impact_csv=args.crude_impact_csv
    )
    print(f"Wrote dashboard MVP to {args.output}")


if __name__ == "__main__":
    main()
