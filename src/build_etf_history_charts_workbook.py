#!/usr/bin/env python3
"""Build an Excel workbook with one ETF history chart sheet per ticker."""

from __future__ import annotations

import argparse

import pandas as pd
from openpyxl.utils import get_column_letter


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if c in invalid else c for c in name).strip()
    cleaned = cleaned[:31] or "Sheet"
    base = cleaned
    i = 1
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    used.add(cleaned)
    return cleaned


def autofit_worksheet(
    ws,
    *,
    min_width: float = 10.0,
    max_width: float = 80.0,
    padding: float = 2.0,
) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                continue
            max_len = max(max_len, len(str(value)))
        width = max(min_width, min(max_width, max_len + padding))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_chart_workbook(etf_csv: str, output_xlsx: str) -> None:
    from openpyxl.chart import LineChart, Reference

    frame = pd.read_csv(etf_csv)
    frame["Date"] = pd.to_datetime(frame["Date"], errors="coerce")
    frame = frame.dropna(subset=["Date"]).sort_values("Date")

    price_cols = [c for c in frame.columns if c.endswith(" - Close")]
    if not price_cols:
        raise RuntimeError("No '* - Close' columns found in ETF CSV.")

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        # Include a compact index sheet.
        index_rows = []
        for col in price_cols:
            parts = col.split(" - ")
            label = parts[0] if len(parts) >= 3 else col
            ticker = parts[1] if len(parts) >= 3 else ""
            index_rows.append({"label": label, "ticker": ticker, "column_name": col})
        pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)
        autofit_worksheet(writer.book["Index"])

        wb = writer.book
        used_names = {"Index"}
        for col in price_cols:
            parts = col.split(" - ")
            label = parts[0] if len(parts) >= 3 else col
            ticker = parts[1] if len(parts) >= 3 else col
            sheet_name = sanitize_sheet_name(f"{ticker}", used_names)

            data = frame[["Date", col]].copy().dropna()
            data.columns = ["Date", "Close"]
            data.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = wb[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            chart = LineChart()
            chart.title = f"{label} / {ticker} Close History"
            chart.y_axis.title = "Close"
            chart.x_axis.title = "Date"
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 8
            chart.width = 14
            ws.add_chart(chart, "D2")

            autofit_worksheet(ws)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Build Excel workbook with one close-price history chart sheet per ETF ticker."
        )
    )
    parser.add_argument("--etf-csv", default="data/outputs/etf_prices.csv")
    parser.add_argument("--output", default="data/outputs/etf_price_history_charts.xlsx")
    args = parser.parse_args()

    build_chart_workbook(args.etf_csv, args.output)
    print(f"Wrote ETF history chart workbook to {args.output}")


if __name__ == "__main__":
    main()
