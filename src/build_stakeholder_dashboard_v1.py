#!/usr/bin/env python3
"""Build stakeholder dashboard v1.0 with MSCI index data.

This script creates a multi-sheet Excel dashboard with:
- Dashboard sheet with dynamic time period selector (1Y/3Y/5Y/10Y)
- Raw_Data sheet with all underlying data
- Control Sheet with column definitions

Data sources:
- MSCI factsheet: Returns (all periods), Market Cap, P/E, factsheet links
- WEO (IMF API): nGDP CAGR (all periods), projections, nGDP 2025 levels
- BIS: REER Index
- Manual (from reference Excel): FX forwards, spot rates
- Manual (WITS/UN): Oil impact data

Output: data/outputs/stakeholder_dashboard_v1.xlsx
"""

import argparse
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation


def load_msci_data(input_path: Path) -> pd.DataFrame:
    """Load MSCI factsheet data with all time periods."""
    df = pd.read_excel(input_path)

    # Exclude aggregate indices (ACWI, WORLD, EM)
    aggregates = {"ACWI", "WORLD", "EM"}
    df = df[~df["country"].isin(aggregates)].copy()

    # Remove duplicates (keep first occurrence)
    df = df.drop_duplicates(subset=["country"], keep="first")

    # Rename columns for clarity
    df = df.rename(columns={
        "country": "country_name",
        "1 yr": "msci_1y",
        "3 yr": "msci_3y",
        "5 yr": "msci_5y",
        "10 yr": "msci_10y",
        "Index Market Cap ($ billion)": "index_market_cap_billion",
        "Top 10 Float Adj Mkt Cap ($ billion)": "top_10_market_cap_billion",
        "P/E": "pe_ratio",
        "link": "factsheet_link"
    })

    return df


def load_weo_data(weo_path: Path) -> pd.DataFrame:
    """Load WEO GDP data and calculate CAGRs for all periods."""
    df = pd.read_csv(weo_path)

    # Filter to NGDPD (nominal GDP in USD) indicator
    ngdp_usd = df[df["indicator"] == "NGDPD"].copy()

    # Pivot to get years as columns
    pivot = ngdp_usd.pivot_table(
        index=["country_name", "country_code"],
        columns="year",
        values="value",
        aggfunc="first"
    ).reset_index()

    # Calculate CAGRs for all periods ending at 2025
    # 1Y: 2025 only (single year, no CAGR, just use the value as-is for consistency)
    # 3Y: 2023-2025
    # 5Y: 2021-2025
    # 10Y: 2016-2025
    
    def calc_cagr(start_year, end_year, row):
        """Calculate CAGR from start to end year."""
        if start_year not in row.index or end_year not in row.index:
            return np.nan
        start_val = row[start_year]
        end_val = row[end_year]
        if pd.isna(start_val) or pd.isna(end_val) or start_val == 0:
            return np.nan
        n = end_year - start_year
        if n == 0:
            return np.nan  # Can't calculate CAGR for single year
        return (end_val / start_val) ** (1/n) - 1

    # Calculate all CAGRs
    # Note: 1Y is not a true CAGR, it's just the latest year's data point
    # We use 2024-2025 for 1Y growth rate
    pivot["ngdp_1y"] = pivot.apply(lambda r: calc_cagr(2024, 2025, r), axis=1)
    pivot["ngdp_3y"] = pivot.apply(lambda r: calc_cagr(2023, 2025, r), axis=1)
    pivot["ngdp_5y"] = pivot.apply(lambda r: calc_cagr(2021, 2025, r), axis=1)
    pivot["ngdp_10y"] = pivot.apply(lambda r: calc_cagr(2016, 2025, r), axis=1)

    # Get 2025 nGDP levels (in billions)
    if 2025 in pivot.columns:
        pivot["ngdp_2025_billion"] = pivot[2025] / 1e9

    # Calculate projected growth 2026-28 CAGR
    if 2026 in pivot.columns and 2028 in pivot.columns:
        pivot["proj_growth_26_28_cagr"] = pivot.apply(lambda r: calc_cagr(2026, 2028, r), axis=1)

    # Select relevant columns
    result = pivot[[
        "country_name", "country_code",
        "ngdp_1y", "ngdp_3y", "ngdp_5y", "ngdp_10y",
        "ngdp_2025_billion", "proj_growth_26_28_cagr"
    ]].copy()

    return result


def load_bis_reer(bis_path: Path) -> pd.DataFrame:
    """Load BIS REER data."""
    df = pd.read_csv(bis_path)

    # Select relevant columns
    result = df[[
        "country_name", "bis_code", "current_reer"
    ]].copy()

    # Calculate implied currency effect from REER
    result["reer_implied_fx_effect"] = (result["current_reer"] - 100) / 100

    return result


def load_fx_manual(fx_path: Path) -> pd.DataFrame:
    """Load manual FX forwards and spot rates."""
    df = pd.read_csv(fx_path)

    # Rename for consistency
    df = df.rename(columns={
        "country": "country_name",
        "forward_rate_1y": "forward_rate_1y",
        "spot_rate": "spot_rate"
    })

    return df


def load_oil_impact(oil_path: Path) -> pd.DataFrame:
    """Load oil import impact data."""
    df = pd.read_csv(oil_path)

    # Select relevant columns
    result = df[[
        "country_name", "impact_percent_gdp"
    ]].copy()

    # Classify oil sensitivity
    def classify_sensitivity(impact):
        if pd.isna(impact):
            return ""  # Blank for missing data
        if impact < 0.2:
            return "low"
        elif impact < 0.5:
            return "med"  # Abbreviated
        else:
            return "high"

    result["oil_sensitivity"] = result["impact_percent_gdp"].apply(classify_sensitivity)

    return result


def build_raw_data(
    msci: pd.DataFrame,
    weo: pd.DataFrame,
    bis: pd.DataFrame,
    fx: pd.DataFrame,
    oil: pd.DataFrame
) -> pd.DataFrame:
    """Build Raw_Data sheet with all time periods and static data."""
    
    # Start with MSCI as base
    df = msci.copy()
    
    # Merge all datasets
    df = df.merge(weo, on="country_name", how="left")
    df = df.merge(bis, on="country_name", how="left")
    df = df.merge(fx, on="country_name", how="left")
    df = df.merge(oil, on="country_name", how="left")
    
    # FILTER 1: Remove countries with no MSCI return data (all periods blank)
    msci_cols = ["msci_1y", "msci_3y", "msci_5y", "msci_10y"]
    df = df[df[msci_cols].notna().any(axis=1)].copy()
    print(f"  After MSCI filter: {len(df)} countries (removed countries with no MSCI data)")
    
    # FILTER 2: Remove countries with Index Market Cap < $25 billion
    df = df[df["index_market_cap_billion"] >= 25].copy()
    print(f"  After Market Cap filter (>= $25B): {len(df)} countries")
    
    # Select and order columns for Raw_Data sheet
    raw_columns = [
        "country_name",
        "msci_1y", "msci_3y", "msci_5y", "msci_10y",
        "ngdp_1y", "ngdp_3y", "ngdp_5y", "ngdp_10y",
        "current_reer", "reer_implied_fx_effect",
        "forward_rate_1y", "spot_rate",
        "impact_percent_gdp", "oil_sensitivity",
        "ngdp_2025_billion",
        "proj_growth_26_28_cagr",
        "index_market_cap_billion", "top_10_market_cap_billion",
        "pe_ratio", "factsheet_link"
    ]
    
    return df[raw_columns].copy()


def write_control_sheet(ws):
    """Write Control Sheet with column definitions."""
    from openpyxl.styles import Font
    
    # Headers
    headers = ["Column Name", "Definition", "Data Source", "Data as of", "Source Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Column definitions
    definitions = [
        ("nGDP (USD) CAGR (%)", "Annualized growth rate of country's nominal GDP in USD terms over selected period", "WEO Database, IMF", "Oct 2025", "https://data.imf.org/en/Data-Explorer"),
        ("MSCI Index Returns (USD) CAGR (%)", "Annualized return of the MSCI country equity index over a selected period.", "MSCI Factsheets", "Feb 27, 2026", "https://www.msci.com/documents/10199/255599/msci-usa-index-gross.pdf"),
        ("Macro Gap (%)", "Difference between a country's GDP growth (USD) and its index return (USD) over the same period.", "", "", ""),
        ("Projected Growth (26-28) CAGR (%)", "IMF forecast of annualized nominal GDP growth in USD terms for 2026-2028.", "WEO Database, IMF", "Oct 2025", "https://data.imf.org/en/Data-Explorer"),
        ("REER Index", "Latest level of the BIS Real Effective Exchange Rate (base year 2020=100). Shows currency change since 2020.", "Bank of International Settlements", "Feb 2026", "https://data.bis.org/topics/EER/data?data_view=tab"),
        ("Implied currency effect from REER", "Estimated percentage overvaluation (+) or undervaluation (-) based on REER deviation from 100. Simplified heuristic.", "", "", ""),
        ("Forward rate for USD/LCU 1 year out (% change)", "Market-implied forecast of how much the Local Currency will move against USD over next 12 months.", "FXEmpire", "2026-03-19", ""),
        ("Avg Currency signals", "Composite indicator averaging multiple currency valuation signals (REER implied + forward rate).", "", "", ""),
        ("Currency Forecast", "Classification of expected currency direction based on Avg Currency Signals.", "", "", ""),
        ("Impact of $10 Oil Price rise relative to GDP (%)", "Economic sensitivity metric: value of $10/barrel oil price increase as % of GDP.", "WITS, World Bank", "2024", "https://wits.worldbank.org/trade/comtrade/en/country/all"),
        ("Oil Sensitivity", "Categorical classification of economic vulnerability to oil price shocks (low/medium/high).", "", "", ""),
        ("nGDP 2025 ($ billion)", "Nominal GDP level in 2025, in billions of USD.", "WEO Database, IMF", "Oct 2025", "https://data.imf.org/en/Data-Explorer"),
        ("USD/LCU spot rate as of Mar 19, 2026", "Current spot exchange rate: how many Local Currency Units per 1 USD.", "FXEmpire", "2026-03-19", ""),
        ("Futures rate for USD/LCU (1 years out from Mar 19, 2026)", "1-year forward exchange rate: market-implied future LCU per USD.", "FXEmpire", "2026-03-19", ""),
        ("Index Market Cap ($ billion)", "Total market capitalization of MSCI country index, in billions USD.", "MSCI Factsheets", "Feb 27, 2026", ""),
        ("Top 10 Float Adj Mkt Cap ($ billion)", "Market cap of top 10 holdings in MSCI country index, in billions USD.", "MSCI Factsheets", "Feb 27, 2026", ""),
        ("P/E", "Price-to-earnings ratio of MSCI country index.", "MSCI Factsheets", "Feb 27, 2026", ""),
        ("MSCI Factsheet Link", "URL to MSCI country index factsheet PDF.", "MSCI Factsheets", "Feb 27, 2026", ""),
    ]
    
    for row, (col_name, definition, source, data_as_of, link) in enumerate(definitions, 2):
        ws.cell(row=row, column=1, value=col_name)
        ws.cell(row=row, column=2, value=definition)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=4, value=data_as_of)
        ws.cell(row=row, column=5, value=link)
    
    # Set column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 124
    ws.column_dimensions["C"].width = 31
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 48


def build_dashboard(
    msci_path: Path,
    weo_path: Path,
    bis_path: Path,
    fx_path: Path,
    oil_path: Path,
    output_path: Path
):
    """Build the complete stakeholder dashboard with dynamic time period selector."""

    print("Loading data sources...")

    # Load all data
    msci = load_msci_data(msci_path)
    print(f"  MSCI: {len(msci)} countries")

    weo = load_weo_data(weo_path)
    print(f"  WEO: {len(weo)} countries")

    bis = load_bis_reer(bis_path)
    print(f"  BIS REER: {len(bis)} countries")

    fx = load_fx_manual(fx_path)
    print(f"  FX Manual: {len(fx)} countries")

    oil = load_oil_impact(oil_path)
    print(f"  Oil Impact: {len(oil)} countries")

    # Build Raw_Data sheet
    print("\nBuilding Raw_Data sheet...")
    raw_data = build_raw_data(msci, weo, bis, fx, oil)
    print(f"  Raw_Data: {len(raw_data)} rows × {len(raw_data.columns)} columns")

    # Write to Excel
    print(f"\nWriting to {output_path}...")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write Raw_Data sheet first
        raw_data.to_excel(writer, sheet_name="Raw_Data", index=False, startrow=0)
        
        # Write Dashboard sheet (empty dataframe, we'll write formulas manually)
        dashboard_df = pd.DataFrame(columns=[
            "Country", "nGDP (USD) CAGR (%)", "MSCI Index Return (USD) CAGR (%)",
            "Macro Gap (%)", "Projected Growth (2026-28) CAGR (%)",
            "REER Index", "Implied currency effect from REER",
            "Forward rate for USD/LCU 1 year out (% change)",
            "Avg Currency signals", "Currency Forecast",
            "Impact of $10 Oil Price rise relative to GDP (%)", "Oil Sensitivity",
            "",  # Spacer column M
            "nGDP 2025 ($ billion)", "USD/LCU spot rate as of Mar 19, 2026",
            "Futures rate for USD/LCU (1 years out from Mar 19, 2026)",
            "Index Market Cap ($ billion)", "Top 10 Float Adj Mkt Cap ($ billion)",
            "P/E", "MSCI Factsheet Link"
        ])
        
        # Add country names
        dashboard_df["Country"] = raw_data["country_name"]
        
        # Write Dashboard sheet starting at row 5 (data rows), headers written manually to row 4
        # Use header=False since we write headers manually
        # startrow=4 means Excel row 5 (0-indexed)
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=4, header=False)
        
        # Get workbook and worksheets
        workbook = writer.book
        ws_dashboard = writer.sheets["Dashboard"]
        ws_raw = writer.sheets["Raw_Data"]
        
        # Create Control Sheet
        ws_control = workbook.create_sheet("Control Sheet")
        write_control_sheet(ws_control)
        
        # === FORMAT DASHBOARD SHEET ===
        
        # Row 1: Title
        header_date = "Feb 27, 2026"  # Fixed to match MSCI factsheet date
        ws_dashboard.cell(row=1, column=1, value=f"Country Disconnect Dashboard (As of {header_date})")
        
        # Row 2: Time Period selector label
        ws_dashboard.cell(row=2, column=1, value="Time Period =")
        
        # Row 3: Time Period dropdown (1Y, 3Y, 5Y, 10Y)
        ws_dashboard.cell(row=3, column=1, value="10Y")
        
        # Add dropdown validation to A3
        dv = DataValidation(type="list", formula1='"1Y,3Y,5Y,10Y"', allow_blank=False)
        dv.error = "Please select a valid time period"
        dv.errorTitle = "Invalid Time Period"
        ws_dashboard.add_data_validation(dv)
        dv.add(ws_dashboard.cell(row=3, column=1))
        
        # Row 4: Column headers (write manually)
        headers = [
            "Country", "nGDP (USD) CAGR (%)", "MSCI Index Return (USD) CAGR (%)",
            "Macro Gap (%)", "Projected Growth (2026-28) CAGR (%)",
            "REER Index", "Implied currency effect from REER",
            "Forward rate for USD/LCU 1 year out (% change)",
            "Avg Currency signals", "Currency Forecast",
            "Impact of $10 Oil Price rise relative to GDP (%)", "Oil Sensitivity",
            "",  # Column M: GAP (visible, narrow separator)
            "",  # Column N: Spacer (hidden)
            "nGDP 2025 ($ billion)", "USD/LCU spot rate as of Mar 19, 2026",
            "Futures rate for USD/LCU (1 years out from Mar 19, 2026)",
            "Index Market Cap ($ billion)", "Top 10 Float Adj Mkt Cap ($ billion)",
            "P/E", "MSCI Factsheet Link"
        ]
        for col, header in enumerate(headers, 1):
            ws_dashboard.cell(row=4, column=col, value=header)
        
        # === WRITE FORMULAS FOR DYNAMIC COLUMNS (B, C, D) ===
        # These formulas reference Raw_Data sheet and use the time period selector in A3
        # Note: Column A (Country) is already populated by dashboard_df
        
        num_countries = len(raw_data)
        first_data_row = 5  # Row 5 is first country (headers on row 4)
        last_data_row = first_data_row + num_countries - 1
        
        for row in range(first_data_row, last_data_row + 1):
            country_row = row - first_data_row + 2  # Raw_Data row (1-indexed, +2 for header)
            
            # Column B: nGDP (USD) CAGR (%) - dynamic based on A3 selector
            # Formula: =INDEX(Raw_Data!$F:$I, MATCH($A5, Raw_Data!$A:$A, 0), MATCH($A$3, {"1Y","3Y","5Y","10Y"}, 0))
            # Raw_Data columns: F=ngdp_1y, G=ngdp_3y, H=ngdp_5y, I=ngdp_10y
            ws_dashboard.cell(row=row, column=2, value=f'=IFERROR(INDEX(Raw_Data!$F:$I, MATCH($A{row}, Raw_Data!$A:$A, 0), MATCH($A$3, {{"1Y","3Y","5Y","10Y"}}, 0)), NA())')
            
            # Column C: MSCI Index Return (USD) CAGR (%) - dynamic based on A3 selector
            # Raw_Data columns: B=msci_1y, C=msci_3y, D=msci_5y, E=msci_10y
            ws_dashboard.cell(row=row, column=3, value=f'=IFERROR(INDEX(Raw_Data!$B:$E, MATCH($A{row}, Raw_Data!$A:$A, 0), MATCH($A$3, {{"1Y","3Y","5Y","10Y"}}, 0)), NA())')
            
            # Column D: Macro Gap (%) = B - C
            ws_dashboard.cell(row=row, column=4, value=f'=IFERROR(B{row}-C{row}, NA())')

            # Column E: Projected Growth (static) - Raw_Data col Q - show NA if blank or zero
            ws_dashboard.cell(row=row, column=5, value=f'=IFERROR(IF(OR(INDEX(Raw_Data!$Q:$Q, MATCH($A{row}, Raw_Data!$A:$A, 0))=\"\", INDEX(Raw_Data!$Q:$Q, MATCH($A{row}, Raw_Data!$A:$A, 0))=0), NA(), INDEX(Raw_Data!$Q:$Q, MATCH($A{row}, Raw_Data!$A:$A, 0))), NA())')

            # Column F: REER Index (static) - Raw_Data col J - show NA if blank
            ws_dashboard.cell(row=row, column=6, value=f'=IFERROR(IF(OR(INDEX(Raw_Data!$J:$J, MATCH($A{row}, Raw_Data!$A:$A, 0))=\"\", INDEX(Raw_Data!$J:$J, MATCH($A{row}, Raw_Data!$A:$A, 0))=0), NA(), INDEX(Raw_Data!$J:$J, MATCH($A{row}, Raw_Data!$A:$A, 0))), NA())')

            # Column G: Implied currency effect from REER = (100 - F) / 100 (will be NA if F is NA)
            ws_dashboard.cell(row=row, column=7, value=f'=IFERROR((100-F{row})/100, NA())')

            # Column H: Forward rate (static) - Raw_Data col L
            ws_dashboard.cell(row=row, column=8, value=f'=IFERROR(INDEX(Raw_Data!$L:$L, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column I: Avg Currency signals = (G + H) / 2 (will be NA if G or H is NA)
            ws_dashboard.cell(row=row, column=9, value=f'=IFERROR((G{row}+H{row})/2, NA())')

            # Column J: Currency Forecast (formula - will show blank if I is NA)
            ws_dashboard.cell(row=row, column=10, value=f'=IF(ISNA(I{row}), \"\", IF(I{row}>0.05,\"undervalued\", IF(I{row}<-0.05,\"overvalued\", \"neutral\")))')

            # Column K: Oil Impact (static) - Raw_Data col N - show NA if blank or zero
            ws_dashboard.cell(row=row, column=11, value=f'=IFERROR(IF(OR(INDEX(Raw_Data!$N:$N, MATCH($A{row}, Raw_Data!$A:$A, 0))=\"\", INDEX(Raw_Data!$N:$N, MATCH($A{row}, Raw_Data!$A:$A, 0))=0), NA(), INDEX(Raw_Data!$N:$N, MATCH($A{row}, Raw_Data!$A:$A, 0))), NA())')

            # Column L: Oil Sensitivity (static) - Raw_Data col O
            ws_dashboard.cell(row=row, column=12, value=f'=IFERROR(INDEX(Raw_Data!$O:$O, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column M: GAP (visible, narrow separator) - blank
            ws_dashboard.cell(row=row, column=13, value='')

            # Column N: Spacer (hidden) - blank
            ws_dashboard.cell(row=row, column=14, value='')

            # Column O: nGDP 2025 ($ billion) - Raw_Data col P
            ws_dashboard.cell(row=row, column=15, value=f'=IFERROR(INDEX(Raw_Data!$P:$P, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column P: USD/LCU spot rate - Raw_Data col M
            ws_dashboard.cell(row=row, column=16, value=f'=IFERROR(INDEX(Raw_Data!$M:$M, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column Q: Futures rate (same as forward rate)
            ws_dashboard.cell(row=row, column=17, value=f'=H{row}')

            # Column R: Index Market Cap ($ billion) - Raw_Data col R
            ws_dashboard.cell(row=row, column=18, value=f'=IFERROR(INDEX(Raw_Data!$R:$R, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column S: Top 10 Market Cap ($ billion) - Raw_Data col S
            ws_dashboard.cell(row=row, column=19, value=f'=IFERROR(INDEX(Raw_Data!$S:$S, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column T: P/E - Raw_Data col T
            ws_dashboard.cell(row=row, column=20, value=f'=IFERROR(INDEX(Raw_Data!$T:$T, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')

            # Column U: MSCI Factsheet Link - Raw_Data col U
            ws_dashboard.cell(row=row, column=21, value=f'=IFERROR(INDEX(Raw_Data!$U:$U, MATCH($A{row}, Raw_Data!$A:$A, 0)), NA())')
        
        # === APPLY NUMBER FORMATTING ===
        for row in range(first_data_row, last_data_row + 1):
            # Percentage columns (0.0%)
            for col in [2, 3, 4, 5, 7, 8, 9]:  # B, C, D, E, G, H, I
                ws_dashboard.cell(row=row, column=col).number_format = '0.0%'
            
            # Decimal columns (0.0) - for billions and indices
            for col in [6, 11, 15, 16]:  # F (REER), K (Oil Impact), O (nGDP 2025), P (Spot rate)
                ws_dashboard.cell(row=row, column=col).number_format = '0.0'
            
            # Decimal columns (0.00) - for market cap in billions
            for col in [18, 19]:  # R (Index Market Cap), S (Top 10 Market Cap)
                ws_dashboard.cell(row=row, column=col).number_format = '0.00'
            
            # P/E ratio (0.0)
            ws_dashboard.cell(row=row, column=20).number_format = '0.0'
            
            # Text/General columns
            for col in [10, 12, 13, 14, 17, 21]:  # J, L, M, N, Q, U
                ws_dashboard.cell(row=row, column=col).number_format = 'General'
        
        # Hide ONLY spacer column N (column M is visible gap)
        ws_dashboard.column_dimensions['M'].width = 2.5  # Narrow visible gap
        ws_dashboard.column_dimensions['N'].hidden = True
        
        # Set Dashboard as active/default sheet
        workbook.active = 1  # Dashboard is sheet index 1 (after Raw_Data at 0)
        
        # === APPLY CONDITIONAL FORMATTING ===
        green_light = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        green_saturated = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        red_light = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        red_saturated = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        yellow_light = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Green/Red scale for B:E (GDP, MSCI, Macro Gap, Projected)
        for col in range(2, 6):  # B, C, D, E
            col_letter = chr(64 + col)
            range_str = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
            ws_dashboard.conditional_formatting.add(range_str, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_light))
            ws_dashboard.conditional_formatting.add(range_str, CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=green_saturated))
            ws_dashboard.conditional_formatting.add(range_str, CellIsRule(operator="lessThan", formula=["0"], fill=red_light))
        
        # Green/Red scale for I (Avg Currency signals)
        range_i = f"I{first_data_row}:I{last_data_row}"
        ws_dashboard.conditional_formatting.add(range_i, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_light))
        ws_dashboard.conditional_formatting.add(range_i, CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=green_saturated))
        ws_dashboard.conditional_formatting.add(range_i, CellIsRule(operator="lessThan", formula=["0"], fill=red_light))
        
        # Oil Impact (K) - lower is better
        range_k = f"K{first_data_row}:K{last_data_row}"
        ws_dashboard.conditional_formatting.add(range_k, CellIsRule(operator="lessThanOrEqual", formula=["0.2"], fill=green_saturated))
        ws_dashboard.conditional_formatting.add(range_k, CellIsRule(operator="lessThanOrEqual", formula=["0.5"], fill=green_light))
        ws_dashboard.conditional_formatting.add(range_k, CellIsRule(operator="greaterThan", formula=["0.5"], fill=red_light))
        
        # Oil Sensitivity (L) - categorical
        # Note: Conditional formatting for text values requires different approach
        # For now, skip this as the values come from Raw_Data
        
        # === SET COLUMN WIDTHS (from reference file) ===
        column_widths = {
            'A': 15.93, 'B': 15.3, 'C': 22.51, 'D': 12.26, 'E': 22.89,
            'F': 12.0, 'G': 22.89, 'H': 34.14, 'I': 19.09, 'J': 14.0,
            'K': 29.47, 'L': 13.28, 'N': 4.67, 'O': 14.0,  # M is visible gap (set separately)
            'P': 27.19, 'Q': 34.4, 'R': 22.38, 'S': 27.19, 'T': 6.7, 'U': 40.0
        }
        for col_letter, width in column_widths.items():
            ws_dashboard.column_dimensions[col_letter].width = width

        # Hide ONLY spacer column N (column M is visible gap)
        ws_dashboard.column_dimensions['N'].hidden = True
        
        # Set autofilter on header row (row 4)
        ws_dashboard.auto_filter.ref = f"A4:U{last_data_row}"
        
        # Freeze panes: header row and selector rows
        ws_dashboard.freeze_panes = "A5"
        
        # Format Raw_Data sheet
        ws_raw.auto_filter.ref = ws_raw.dimensions
        ws_raw.freeze_panes = "A2"
        
        # Format Control Sheet
        ws_control.auto_filter.ref = "A1:E20"
        ws_control.freeze_panes = "A2"

    print(f"\nDashboard created successfully with {num_countries} countries!")
    print(f"Output: {output_path}")
    print(f"Sheets: Dashboard, Raw_Data, Control Sheet")


def main():
    parser = argparse.ArgumentParser(
        description="Build stakeholder dashboard v1.0 with MSCI index data"
    )
    parser.add_argument(
        "--msci",
        type=Path,
        default=Path("data/inputs/factsheet data manual_feb 27 2026.xlsx"),
        help="Path to MSCI factsheet Excel file"
    )
    parser.add_argument(
        "--weo",
        type=Path,
        default=Path("data/outputs/weo_gdp.csv"),
        help="Path to WEO GDP CSV file"
    )
    parser.add_argument(
        "--bis",
        type=Path,
        default=Path("data/outputs/bis_reer_metrics.csv"),
        help="Path to BIS REER CSV file"
    )
    parser.add_argument(
        "--fx",
        type=Path,
        default=Path("data/inputs/fx_forwards_manual.csv"),
        help="Path to manual FX forwards CSV file"
    )
    parser.add_argument(
        "--oil",
        type=Path,
        default=Path("data/outputs/crude_oil_import_impact.csv"),
        help="Path to oil import impact CSV file"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/outputs/stakeholder_dashboard_v1.xlsx"),
        help="Output Excel file path"
    )

    args = parser.parse_args()

    # Resolve paths relative to project root
    script_dir = Path(__file__).parent
    project_root = script_dir.parent

    msci_path = project_root / args.msci
    weo_path = project_root / args.weo
    bis_path = project_root / args.bis
    fx_path = project_root / args.fx
    oil_path = project_root / args.oil
    output_path = project_root / args.output

    build_dashboard(
        msci_path=msci_path,
        weo_path=weo_path,
        bis_path=bis_path,
        fx_path=fx_path,
        oil_path=oil_path,
        output_path=output_path
    )


if __name__ == "__main__":
    main()
